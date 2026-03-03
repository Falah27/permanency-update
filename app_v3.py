"""
MIB Checksheet Generator - Version 3.0 (Phase 2: Enhanced Quality)
Aplikasi web untuk generate checksheet MIB Epson secara otomatis.

Phase 1 Features:
✅ Logging system
✅ Input validation  
✅ Resource cleanup
✅ Error retry mechanism

Phase 2 Features (NEW):
✅ Output verification
✅ Enhanced progress tracking
✅ Metrics collection
✅ Automated testing support
"""

import streamlit as st
import win32com.client as win32
import os
import shutil
import pythoncom
import logging
import time
from datetime import datetime
from functools import wraps

# Import metrics system
try:
    from metrics import get_metrics_collector, record_generation
    METRICS_AVAILABLE = True
except ImportError:
    METRICS_AVAILABLE = False
    logging.warning("Metrics module not available")

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ==========================================
# LOGGING SETUP
# ==========================================

LOG_DIR = "logs"
os.makedirs(LOG_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - [%(funcName)s] %(message)s',
    handlers=[
        logging.FileHandler(f"{LOG_DIR}/app_{datetime.now():%Y%m%d}.log", encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

logger.info("=" * 60)
logger.info("Application started - MIB Checksheet Generator v3.0 (Phase 2)")
logger.info("=" * 60)

# ==========================================
# CONSTANTS & CONFIGURATION
# ==========================================

class ExcelConst:
    """Excel API Constants"""
    CALCULATION_MANUAL = -4135
    CALCULATION_AUTOMATIC = -4105
    SHEET_VISIBLE = -1
    PASTE_ALL = -4163
    HORIZONTAL_CENTER = -4108
    NORMAL_VIEW = 1

class Config:
    """Application Configuration"""
    # Sheet names
    SHEET_PUBLIC_MIB = "Public MIB"
    SHEET_PRIVATE_MIB = "EPSON Private MIB"
    SHEET_DUMP = "dump"
    SHEET_TEMPLATE_SUMMARY = "集計表"
    SHEET_TEMPLATE_OID = "oid"
    
    IMPORTANT_SHEETS = [SHEET_PUBLIC_MIB, SHEET_PRIVATE_MIB, SHEET_DUMP, 
                        SHEET_TEMPLATE_SUMMARY, SHEET_TEMPLATE_OID]
    
    # Column indices
    COL_START_CCODE = 11  # Column K
    COL_END_SCAN = 200
    COL_STEP = 5
    COL_START_EVAL = 17  # Column Q
    
    # Row indices  
    ROW_HEADER = 11
    ROW_SUBHEADER = 12
    ROW_DATA_START = 13
    
    # File & folder
    TEMP_DIR = "temp_workspace"
    TEMPLATE_FILE = "Template_sheet.xlsm"
    
    # Display settings
    DEFAULT_ZOOM = 80
    BATCH_SIZE_UNION = 500
    
    # Cleanup settings
    MAX_FILE_AGE_HOURS = 24
    
    # Retry settings
    MAX_RETRIES = 3
    RETRY_DELAY = 2

# Evaluation column headers
EVAL_HEADERS = ["値(比較用)", "値(比較用） 加工", "取得値", "自動", "手動", 
                "判定理由", "担当者", "自/他", "手動", "判定理由"]

# Excel formulas
FORMULA_SHOURYAKU = '=IF(AND(M13="", O13="", N13<>""), "○", "")'
FORMULA_VALUE_COMPARE = '=IF(P13="","",IF(P13="←",IF(OFFSET($J13,0,MATCH(MID(M13,1,FIND("の",M13,1)-1), $L$11:$BA$11,0)+4,1,1)="","",OFFSET($L13,0,MATCH(MID(M13,1,FIND("の",M13,1)-1), $L$11:$BA$11,0)+4,1,1)),P13))'
FORMULA_VALUE_PROCESS = '=IF(COUNTIF(Q13,"*(*"),MID(Q13,FIND("(",Q13,1)+1,FIND(")",Q13,1)-FIND("(",Q13,1)-1),IF(COUNTIF(Q13,"*""*"),MID(Q13,FIND("""",Q13,1)+1,LEN(Q13)-2),IF(COUNTIF(Q13,"*：*"),RIGHT(Q13,LEN(Q13)-(FIND("：",Q13))),Q13)))'
FORMULA_VLOOKUP = '=IFERROR(IF(VLOOKUP("*"&TRIM($F13)&"*",dump!$A:$D,4,FALSE)="","空文字",VLOOKUP("*"&TRIM($F13)&"*",dump!$A:$D,4,FALSE)),"NA")'
FORMULA_HANTEI = '=IF($K13<>"",IF($E13<>"","■",""),IF(AND(R13="",S13="NA"),"●", IF(EXACT(R13,S13),"●","×")))'

# ==========================================
# DECORATORS
# ==========================================

def measure_time(func):
    """Decorator untuk mengukur waktu eksekusi fungsi"""
    @wraps(func)
    def wrapper(*args, **kwargs):
        start = time.time()
        result = func(*args, **kwargs)
        duration = time.time() - start
        
        logger.info(f"⏱️ {func.__name__} completed in {duration:.2f}s")
        
        if duration > 30:
            logger.warning(f"⚠️ {func.__name__} is slow ({duration:.2f}s)")
        
        return result
    return wrapper

def retry_on_failure(max_retries=None, delay=None):
    """Decorator untuk auto-retry operation yang gagal"""
    if max_retries is None:
        max_retries = Config.MAX_RETRIES
    if delay is None:
        delay = Config.RETRY_DELAY
    
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    if attempt < max_retries - 1:
                        logger.warning(f"🔄 Attempt {attempt+1}/{max_retries} failed: {e}, retrying in {delay}s...")
                        time.sleep(delay)
                    else:
                        logger.error(f"❌ All {max_retries} attempts failed for {func.__name__}")
                        raise
        return wrapper
    return decorator

# ==========================================
# VALIDATION FUNCTIONS
# ==========================================

def validate_spek_file(file_path):
    """
    Validasi file Spek sebelum diproses
    Returns: (is_valid: bool, message: str)
    """
    logger.info(f"Validating Spek file: {file_path}")
    
    try:
        # Cek file exists
        if not os.path.exists(file_path):
            return False, "❌ File tidak ditemukan"
        
        # Cek file size (minimal 100KB, maksimal 50MB)
        file_size = os.path.getsize(file_path) / (1024 * 1024)  # MB
        if file_size < 0.1:
            return False, f"❌ File terlalu kecil ({file_size:.1f}MB), mungkin corrupt"
        if file_size > 50:
            return False, f"❌ File terlalu besar ({file_size:.1f}MB), maksimal 50MB"
        
        # Validasi dengan openpyxl
        if OPENPYXL_AVAILABLE:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            # Cek sheet wajib ada
            if Config.SHEET_PUBLIC_MIB not in wb.sheetnames:
                wb.close()
                return False, f"❌ Sheet '{Config.SHEET_PUBLIC_MIB}' tidak ditemukan"
            
            sheet = wb[Config.SHEET_PUBLIC_MIB]
            
            # Cek struktur header (row 11 harus ada data C-Code di column K-T range)
            # Cek beberapa columns untuk lebih flexible
            has_header_data = False
            for col_idx in range(Config.COL_START_CCODE, Config.COL_START_CCODE + 50, Config.COL_STEP):
                value = sheet.cell(Config.ROW_HEADER, col_idx).value
                if value and str(value).strip():
                    has_header_data = True
                    break
            
            if not has_header_data:
                wb.close()
                return False, "❌ Format file tidak valid (tidak ada C-Code di row 11)"
            
            # Cek minimal ada beberapa rows data (lebih lenient - minimal 20)
            if sheet.max_row < 20:
                wb.close()
                return False, f"❌ Data terlalu sedikit ({sheet.max_row} rows), expected >20"
            
            wb.close()
            logger.info(f"✅ File validation passed (size: {file_size:.1f}MB, rows: {sheet.max_row})")
            return True, f"✅ File valid ({file_size:.1f}MB, {sheet.max_row} rows)"
        else:
            # Fallback: basic file check
            logger.info("✅ Basic validation passed (openpyxl not available)")
            return True, f"✅ File exists ({file_size:.1f}MB)"
    
    except Exception as e:
        logger.error(f"❌ Validation error: {e}", exc_info=True)
        return False, f"❌ File corrupt atau format tidak valid: {str(e)[:100]}"

def validate_template_file():
    """Validasi file template exists"""
    if not os.path.exists(Config.TEMPLATE_FILE):
        logger.warning(f"⚠️ Template file not found: {Config.TEMPLATE_FILE}")
        return False, f"⚠️ Template '{Config.TEMPLATE_FILE}' tidak ditemukan"
    
    logger.info(f"✅ Template file exists: {Config.TEMPLATE_FILE}")
    return True, "✅ Template file ready"

def verify_output_file(output_path):
    """
    Verify generated checksheet output meets quality standards
    Returns: (is_valid: bool, report: Dict, message: str)
    """
    logger.info(f"🔍 Verifying output file: {output_path}")
    
    report = {
        "file_exists": False,
        "file_size_ok": False,
        "sheets_exist": False,
        "formulas_present": False,
        "data_integrity": False,
        "warnings": []
    }
    
    try:
        # Check 1: File exists
        if not os.path.exists(output_path):
            return False, report, "❌ Output file not found"
        report["file_exists"] = True
        
        # Check 2: File size reasonable (> 500KB)
        file_size = os.path.getsize(output_path)
        if file_size < 500 * 1024:
            report["warnings"].append(f"File size too small: {file_size / 1024:.0f}KB")
        else:
            report["file_size_ok"] = True
        
        # Check 3-5: Excel content validation (if openpyxl available)
        if OPENPYXL_AVAILABLE:
            try:
                wb = load_workbook(output_path, data_only=False)
                
                # Check 3: Required sheets exist
                required_sheets = [Config.SHEET_PUBLIC_MIB, Config.SHEET_DUMP]
                sheets_present = all(s in wb.sheetnames for s in required_sheets)
                report["sheets_exist"] = sheets_present
                
                if not sheets_present:
                    missing = [s for s in required_sheets if s not in wb.sheetnames]
                    report["warnings"].append(f"Missing sheets: {', '.join(missing)}")
                
                # Check 4: Formulas present in Public MIB
                if Config.SHEET_PUBLIC_MIB in wb.sheetnames:
                    sheet = wb[Config.SHEET_PUBLIC_MIB]
                    
                    # Check cell T13 for HANTEI formula
                    cell_formula = sheet.cell(Config.ROW_DATA_START, 20).value
                    if cell_formula and ("IF" in str(cell_formula) or "EXACT" in str(cell_formula)):
                        report["formulas_present"] = True
                    else:
                        report["warnings"].append("Formulas may be missing in evaluation columns")
                    
                    # Check 5: Data integrity (sufficient rows)
                    if sheet.max_row >= 100:
                        report["data_integrity"] = True
                    else:
                        report["warnings"].append(f"Low row count: {sheet.max_row}")
                
                wb.close()
                
            except Exception as e:
                logger.warning(f"Excel content validation failed: {e}")
                report["warnings"].append(f"Content check error: {str(e)[:50]}")
        else:
            # Can't verify Excel content without openpyxl
            report["warnings"].append("openpyxl not available - limited verification")
        
        # Overall validation
        critical_checks = [
            report["file_exists"],
            report["file_size_ok"],
            report["sheets_exist"]
        ]
        
        is_valid = all(critical_checks)
        
        if is_valid:
            if report["warnings"]:
                msg = f"⚠️ Valid with warnings: {len(report['warnings'])} issues"
            else:
                msg = "✅ Output verified successfully"
        else:
            failed = [k for k, v in report.items() if k != "warnings" and not v]
            msg = f"❌ Verification failed: {', '.join(failed)}"
        
        logger.info(f"Verification result: {msg}")
        return is_valid, report, msg
        
    except Exception as e:
        logger.error(f"Output verification error: {e}", exc_info=True)
        return False, report, f"❌ Verification error: {str(e)[:100]}"

# ==========================================
# RESOURCE CLEANUP
# ==========================================

def cleanup_old_temp_files(max_age_hours=None):
    """
    Hapus file temporary yang sudah lama
    Returns: jumlah file yang dihapus
    """
    if max_age_hours is None:
        max_age_hours = Config.MAX_FILE_AGE_HOURS
    
    logger.info(f"🧹 Starting cleanup (max age: {max_age_hours}h)")
    
    try:
        if not os.path.exists(Config.TEMP_DIR):
            logger.info("Temp directory doesn't exist, skipping cleanup")
            return 0
        
        deleted_count = 0
        current_time = time.time()
        
        for filename in os.listdir(Config.TEMP_DIR):
            file_path = os.path.join(Config.TEMP_DIR, filename)
            
            # Skip if not a file
            if not os.path.isfile(file_path):
                continue
            
            # Cek umur file
            file_age_hours = (current_time - os.path.getmtime(file_path)) / 3600
            
            if file_age_hours > max_age_hours:
                try:
                    os.remove(file_path)
                    deleted_count += 1
                    logger.info(f"🗑️ Deleted old file: {filename} (age: {file_age_hours:.1f}h)")
                except Exception as e:
                    logger.error(f"Failed to delete {filename}: {e}")
        
        logger.info(f"✅ Cleanup completed: {deleted_count} files deleted")
        return deleted_count
    
    except Exception as e:
        logger.error(f"❌ Cleanup failed: {e}", exc_info=True)
        return 0

# ==========================================
# EXCEL CORE OPERATIONS
# ==========================================

@retry_on_failure(max_retries=2, delay=1)
def init_excel_app():
    """Initialize Excel application in background mode."""
    logger.info("🔌 Initializing Excel COM application")
    pythoncom.CoInitialize()
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    excel.EnableEvents = False
    logger.info("✅ Excel initialized successfully")
    return excel

def close_excel_safely(excel, wb=None):
    """Close Excel application and cleanup COM."""
    logger.info("🔌 Closing Excel application")
    try:
        workbook_count = excel.Workbooks.Count
        while excel.Workbooks.Count > 0:
            excel.Workbooks(1).Close(SaveChanges=False)
        logger.info(f"Closed {workbook_count} workbook(s)")
    except Exception as e:
        logger.error(f"Error closing workbooks: {e}")
    
    try:
        excel.ScreenUpdating = True
        excel.EnableEvents = True
        excel.DisplayAlerts = True
        excel.Quit()
        logger.info("✅ Excel closed successfully")
    except Exception as e:
        logger.error(f"Error quitting Excel: {e}")
    
    try:
        pythoncom.CoUninitialize()
    except Exception as e:
        logger.error(f"Error uninitializing COM: {e}")

def find_last_row(sheet):
    """Find the last used row in a sheet."""
    try:
        if sheet.UsedRange.Rows.Count > 0:
            last_row = sheet.UsedRange.Rows.Count + sheet.UsedRange.Row - 1
            if last_row > 10:
                logger.debug(f"Last row found (UsedRange): {last_row}")
                return last_row
        
        last_cell = sheet.Cells.Find(What="*", SearchOrder=1, SearchDirection=2)
        if last_cell and last_cell.Row > 10:
            logger.debug(f"Last row found (Find): {last_cell.Row}")
            return last_cell.Row
    except Exception as e:
        logger.warning(f"Error finding last row: {e}")
    
    logger.warning("Using fallback last row: 500")
    return 500

# ==========================================
# C-CODE DETECTION
# ==========================================

@measure_time
def read_ccodes_openpyxl(file_path):
    """Fast C-Code detection using openpyxl (no Excel needed)."""
    logger.info(f"📊 Reading C-Codes with openpyxl from: {file_path}")
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb[Config.SHEET_PUBLIC_MIB]
    ccodes = []
    
    for col_idx in range(Config.COL_START_CCODE, Config.COL_END_SCAN, Config.COL_STEP):
        value = sheet.cell(row=Config.ROW_HEADER, column=col_idx).value
        if not value or str(value).strip() == "":
            break
        ccodes.append({"name": str(value).strip(), "col_index": col_idx})
    
    wb.close()
    logger.info(f"✅ Detected {len(ccodes)} C-Codes: {[c['name'] for c in ccodes]}")
    return ccodes

@measure_time
def read_ccodes_com(wb):
    """Fallback C-Code detection using Excel COM (batch read)."""
    logger.info("📊 Reading C-Codes with COM (fallback mode)")
    sheet = wb.Sheets(Config.SHEET_PUBLIC_MIB)
    ccodes = []
    
    data_range = sheet.Range(
        sheet.Cells(Config.ROW_HEADER, Config.COL_START_CCODE),
        sheet.Cells(Config.ROW_HEADER, Config.COL_END_SCAN)
    ).Value
    
    if data_range:
        if not isinstance(data_range, tuple):
            data_range = (data_range,)
        elif isinstance(data_range[0], tuple):
            data_range = data_range[0]
        
        for col_idx in range(Config.COL_START_CCODE, Config.COL_END_SCAN + 1, Config.COL_STEP):
            array_idx = col_idx - Config.COL_START_CCODE
            if array_idx >= len(data_range):
                break
            
            value = data_range[array_idx]
            if not value or str(value).strip() == "":
                break
            
            ccodes.append({"name": str(value).strip(), "col_index": col_idx})
    
    logger.info(f"✅ Detected {len(ccodes)} C-Codes")
    return ccodes

# ==========================================
# SHEET OPERATIONS
# ==========================================

def create_dump_sheet(wb):
    """Create empty dump sheet for VLOOKUP."""
    logger.info("📝 Creating dump sheet")
    try:
        target_sheet = wb.Sheets(Config.SHEET_PRIVATE_MIB)
        new_sheet = wb.Sheets.Add(After=target_sheet)
    except:
        new_sheet = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
    
    new_sheet.Name = Config.SHEET_DUMP
    logger.info("✅ Dump sheet created")

def insert_shouryaku_column(sheet, last_row):
    """Insert 省略 column."""
    logger.debug("Inserting 省略 column")
    sheet.Columns(Config.COL_START_CCODE).Insert()
    
    header_cell = sheet.Cells(Config.ROW_HEADER, Config.COL_START_CCODE)
    header_cell.Value = "省略"
    header_cell.Font.Bold = False
    header_cell.Font.Size = 9
    
    merge_area = sheet.Range(
        sheet.Cells(9, Config.COL_START_CCODE),
        sheet.Cells(Config.ROW_SUBHEADER, Config.COL_START_CCODE)
    )
    merge_area.Merge()
    
    format_area = sheet.Range(
        sheet.Cells(9, Config.COL_START_CCODE),
        sheet.Cells(last_row, Config.COL_START_CCODE)
    )
    format_area.Interior.ColorIndex = 2
    format_area.Borders.LineStyle = 1
    format_area.Borders.Weight = 2
    format_area.HorizontalAlignment = ExcelConst.HORIZONTAL_CENTER

def delete_unused_columns(sheet, target_col):
    """Delete columns outside the target C-Code."""
    logger.debug(f"Deleting unused columns (keeping col {target_col})")
    new_target = target_col + 1
    delete_start = new_target + 5
    
    sheet.Range(
        sheet.Columns(delete_start),
        sheet.Columns(Config.COL_END_SCAN)
    ).Delete()
    
    if new_target > 12:
        sheet.Range(
            sheet.Columns(12),
            sheet.Columns(new_target - 1)
        ).Delete()

@measure_time
def add_evaluation_columns(sheet, last_row, excel_app):
    """Add evaluation columns with formulas."""
    logger.info("📊 Adding evaluation columns")
    start_col = Config.COL_START_EVAL
    
    original_calc = excel_app.Calculation
    excel_app.Calculation = ExcelConst.CALCULATION_MANUAL
    
    try:
        clear_area = sheet.Range(
            sheet.Cells(1, start_col),
            sheet.Cells(min(last_row + 10, 5000), start_col + 9)
        )
        clear_area.ClearFormats()
        
        header_range = sheet.Range(
            sheet.Cells(Config.ROW_SUBHEADER, start_col),
            sheet.Cells(Config.ROW_SUBHEADER, start_col + len(EVAL_HEADERS) - 1)
        )
        header_range.Value = tuple(EVAL_HEADERS)
        header_range.Font.Bold = False
        
        table_area = sheet.Range(
            sheet.Cells(9, start_col),
            sheet.Cells(last_row, start_col + 9)
        )
        table_area.Interior.ColorIndex = 2
        table_area.Borders.LineStyle = 1
        table_area.Borders.Weight = 2
        table_area.ColumnWidth = 12
        table_area.WrapText = True
        table_area.Font.Size = 9
        
        if last_row >= Config.ROW_DATA_START:
            _apply_evaluation_formulas(sheet, last_row)
            _clear_nw_rows(sheet, last_row)
        
        logger.info("✅ Evaluation columns added")
    finally:
        excel_app.Calculation = original_calc

def _apply_evaluation_formulas(sheet, last_row):
    """Apply evaluation formulas to ranges."""
    logger.debug("Applying evaluation formulas")
    
    area_k = sheet.Range(
        sheet.Cells(Config.ROW_DATA_START, 11),
        sheet.Cells(last_row, 11)
    )
    area_k.Formula = FORMULA_SHOURYAKU
    
    areas = [
        (17, FORMULA_VALUE_COMPARE),
        (18, FORMULA_VALUE_PROCESS),
        (19, FORMULA_VLOOKUP),
        (20, FORMULA_HANTEI)
    ]
    
    for col_idx, formula in areas:
        area = sheet.Range(
            sheet.Cells(Config.ROW_DATA_START, col_idx),
            sheet.Cells(last_row, col_idx)
        )
        area.Formula = formula

def _clear_nw_rows(sheet, last_row):
    """Clear evaluation columns for NW rows."""
    logger.debug("Clearing NW rows")
    data_j = sheet.Range(
        sheet.Cells(Config.ROW_DATA_START, 10),
        sheet.Cells(last_row, 10)
    ).Value
    
    if not data_j:
        return
    
    ranges_to_clear = []
    for i, row_data in enumerate(data_j):
        value = str(row_data[0]) if row_data[0] else ""
        if "NW" in value:
            row_num = Config.ROW_DATA_START + i
            ranges_to_clear.append(f"Q{row_num}:Z{row_num}")
    
    if ranges_to_clear:
        logger.debug(f"Clearing {len(ranges_to_clear)} NW rows")
        for batch_start in range(0, len(ranges_to_clear), Config.BATCH_SIZE_UNION):
            batch = ranges_to_clear[batch_start:batch_start + Config.BATCH_SIZE_UNION]
            try:
                combined = sheet.Range(batch[0])
                for addr in batch[1:]:
                    try:
                        combined = sheet.Application.Union(combined, sheet.Range(addr))
                    except:
                        pass
                combined.ClearContents()
            except:
                pass

def delete_unused_sheets(wb):
    """Delete sheets not in important list."""
    logger.info("🧹 Deleting unused sheets")
    target_lower = [s.lower() for s in Config.IMPORTANT_SHEETS]
    deleted = []
    
    for sheet in wb.Sheets:
        if sheet.Name.lower() not in target_lower:
            try:
                sheet.Delete()
                deleted.append(sheet.Name)
            except:
                pass
    
    if deleted:
        logger.info(f"🗑️ Deleted sheets: {', '.join(deleted)}")
    else:
        logger.info("No sheets to delete")
    
    return deleted

def unhide_all_cells(sheet):
    """Unhide all hidden columns and rows in sheet."""
    logger.debug(f"Unhiding all cells in sheet: {sheet.Name}")
    try:
        sheet.Columns.Hidden = False
        sheet.Rows.Hidden = False
    except Exception as e:
        logger.warning(f"Error unhiding cells: {e}")

# ==========================================
# TEMPLATE INJECTION
# ==========================================

@retry_on_failure(max_retries=2, delay=1)
@measure_time
def inject_template_sheets(wb_target, template_path, excel):
    """Inject template sheets using simplified single-method approach."""
    logger.info(f"📑 Injecting template sheets from: {template_path}")
    
    if not os.path.exists(template_path):
        logger.error(f"Template not found: {template_path}")
        return False, f"❌ Template not found: {template_path}"
    
    wb_template = None
    try:
        wb_template = excel.Workbooks.Open(os.path.abspath(template_path), ReadOnly=True)
        available = [s.Name for s in wb_template.Sheets]
        target_sheets = [Config.SHEET_TEMPLATE_SUMMARY, Config.SHEET_TEMPLATE_OID]
        
        logger.info(f"Available sheets in template: {available}")
        
        copied = []
        errors = []
        
        for target_name in target_sheets:
            source_name = None
            for avail in available:
                if avail.lower() == target_name.lower():
                    source_name = avail
                    break
            
            if not source_name:
                msg = f"Sheet '{target_name}' not found in template"
                logger.warning(msg)
                errors.append(f"⚠️ {msg}")
                continue
            
            try:
                logger.debug(f"Copying sheet: {source_name}")
                source_sheet = wb_template.Sheets(source_name)
                source_sheet.Copy(Before=wb_target.Sheets(1))
                
                new_sheet = wb_target.Sheets(1)
                if new_sheet.Name != target_name:
                    try:
                        new_sheet.Name = target_name
                    except:
                        pass
                
                try:
                    new_sheet.Move(After=wb_target.Sheets(wb_target.Sheets.Count))
                except:
                    logger.debug(f"Move failed for {target_name}, but sheet copied successfully")
                
                copied.append(target_name)
                logger.info(f"✅ Copied sheet: {target_name}")
                
            except Exception as e:
                msg = f"Failed to copy '{target_name}': {str(e)}"
                logger.error(msg)
                errors.append(f"❌ {msg}")
        
        wb_template.Close(SaveChanges=False)
        
        if copied:
            success_msg = f"✅ Copied {len(copied)} sheets: {', '.join(copied)}"
            logger.info(success_msg)
            return True, success_msg
        else:
            error_msg = f"❌ Failed: {'; '.join(errors)}"
            logger.error(error_msg)
            return False, error_msg
    
    except Exception as e:
        logger.error(f"Template injection error: {e}", exc_info=True)
        if wb_template:
            try:
                wb_template.Close(SaveChanges=False)
            except:
                pass
        return False, f"❌ Error: {str(e)}"

# ==========================================
# MAIN PROCESSING
# ==========================================

@measure_time
def process_sheet_cutting(wb, target_col, excel):
    """Main processing: cut columns and add evaluation."""
    logger.info(f"✂️ Processing sheet cutting (target col: {target_col})")
    target_sheets = [Config.SHEET_PUBLIC_MIB, Config.SHEET_PRIVATE_MIB]
    
    original_screen = excel.ScreenUpdating
    original_calc = excel.Calculation
    
    try:
        excel.ScreenUpdating = False
        excel.Calculation = ExcelConst.CALCULATION_MANUAL
        
        for sheet_name in target_sheets:
            logger.info(f"Processing sheet: {sheet_name}")
            sheet = wb.Sheets(sheet_name)
            last_row = find_last_row(sheet)
            logger.info(f"Last row detected: {last_row}")
            
            insert_shouryaku_column(sheet, last_row)
            delete_unused_columns(sheet, target_col)
            add_evaluation_columns(sheet, last_row, excel)
            unhide_all_cells(sheet)
            
            sheet.Activate()
            sheet.PageSetup.PrintArea = ""
        
        logger.info("✅ Sheet cutting completed")
    finally:
        excel.ScreenUpdating = original_screen
        excel.Calculation = original_calc

def finalize_workbook_view(wb, excel):
    """Set final view: unhide sheets, zoom, position."""
    logger.info("🎨 Finalizing workbook view")
    try:
        for sheet in wb.Sheets:
            if any(sheet.Name.lower() == imp.lower() for imp in Config.IMPORTANT_SHEETS):
                sheet.Visible = ExcelConst.SHEET_VISIBLE
        
        main_sheets = [Config.SHEET_PUBLIC_MIB, Config.SHEET_PRIVATE_MIB]
        
        for sheet_name in main_sheets:
            try:
                sheet = wb.Sheets(sheet_name)
                sheet.Activate()
                sheet.Range("K1").Select()
                
                excel.Windows(1).View = ExcelConst.NORMAL_VIEW
                excel.Windows(1).Zoom = Config.DEFAULT_ZOOM
                excel.ActiveWindow.ScrollRow = 1
                excel.ActiveWindow.ScrollColumn = Config.COL_START_CCODE
            except Exception as e:
                logger.warning(f"Error setting view for {sheet_name}: {e}")
        
        wb.Sheets(Config.SHEET_PUBLIC_MIB).Activate()
        logger.info("✅ View finalized")
    
    except Exception as e:
        logger.warning(f"Error finalizing view: {e}")

# ==========================================
# STREAMLIT UI
# ==========================================

def main():
    st.set_page_config(
        page_title="Checksheet Generator v3.0",
        page_icon="🖨️",
        layout="centered"
    )
    
    st.title("🖨️ MIB Checksheet Generator v3.0")
    st.markdown("Aplikasi otomatis untuk generate checksheet MIB Epson.")
    st.caption("✨ Version 3.0 - Phase 2: Enhanced Quality & Metrics")
    st.divider()
    
    # Cleanup old files on startup
    os.makedirs(Config.TEMP_DIR, exist_ok=True)
    deleted = cleanup_old_temp_files()
    if deleted > 0:
        st.info(f"🧹 Cleaned up {deleted} old temporary file(s)")
    
    # === STAGE 1: UPLOAD ===
    st.markdown("### 📥 Stage 1: Upload File")
    
    uploaded_spek = st.file_uploader("📂 Spek Original (.xlsm)", type=["xlsm"])
    
    if uploaded_spek is None:
        st.info("👆 Upload Spek Original untuk mulai")
        
        # Show system status in sidebar
        with st.sidebar:
            st.markdown("### 📊 System Status")
            st.write(f"🔧 openpyxl: {'✅ Available' if OPENPYXL_AVAILABLE else '❌ Not installed'}")
            st.write(f"📁 Temp dir: {Config.TEMP_DIR}")
            st.write(f"📄 Template: {'✅' if os.path.exists(Config.TEMPLATE_FILE) else '❌'} {Config.TEMPLATE_FILE}")
            
            if st.button("🧹 Clean Temp Files Now"):
                deleted = cleanup_old_temp_files(max_age_hours=0)
                st.success(f"Deleted {deleted} file(s)")
        
        return
    
    # Save uploaded file
    spek_path = os.path.join(Config.TEMP_DIR, uploaded_spek.name)
    with open(spek_path, "wb") as f:
        f.write(uploaded_spek.getbuffer())
    
    logger.info(f"📁 File uploaded: {uploaded_spek.name} ({uploaded_spek.size} bytes)")
    
    # === VALIDATION ===
    with st.spinner("🔍 Validating file..."):
        is_valid, validation_msg = validate_spek_file(spek_path)
        
        if not is_valid:
            st.error(validation_msg)
            logger.error(f"Validation failed: {validation_msg}")
            return
        
        st.success(validation_msg)
    
    # === STAGE 2: DETECT C-CODES ===
    cache_key = f"{uploaded_spek.name}_{uploaded_spek.size}"
    
    if "ccodes" not in st.session_state or st.session_state.get('cache_key') != cache_key:
        with st.spinner("⚡ Detecting C-Codes..."):
            try:
                if OPENPYXL_AVAILABLE:
                    st.session_state.ccodes = read_ccodes_openpyxl(spek_path)
                    st.session_state.cache_key = cache_key
                    st.success(f"⚡ Detected {len(st.session_state.ccodes)} C-Codes (fast mode)")
                else:
                    st.info("🐌 Tip: `pip install openpyxl` for 100x faster detection")
                    excel = init_excel_app()
                    try:
                        wb = excel.Workbooks.Open(os.path.abspath(spek_path), ReadOnly=True, UpdateLinks=0)
                        st.session_state.ccodes = read_ccodes_com(wb)
                        st.session_state.cache_key = cache_key
                        st.success(f"✅ Detected {len(st.session_state.ccodes)} C-Codes")
                    finally:
                        close_excel_safely(excel)
            
            except Exception as e:
                st.error(f"❌ Detection failed: {e}")
                logger.error(f"C-Code detection failed: {e}", exc_info=True)
                return
    else:
        st.info(f"📦 Using cache ({len(st.session_state.ccodes)} C-Codes)")
    
    if not st.session_state.ccodes:
        st.warning("⚠️ No C-Codes detected")
        return
    
    # === STAGE 3: GENERATE ===
    st.divider()
    st.markdown("### 🎯 Stage 2: Generate Checksheet")
    
    ccodes_list = [c["name"] for c in st.session_state.ccodes]
    selected_name = st.selectbox("Select C-Code:", ccodes_list)
    
    if st.button("🚀 Generate Checksheet", type="primary", use_container_width=True):
        selected = next(c for c in st.session_state.ccodes if c["name"] == selected_name)
        safe_name = "".join(c for c in selected["name"] if c.isalnum() or c == ' ').strip()
        output_filename = f"Checksheet_{safe_name.replace(' ', '_')}.xlsm"
        output_path = os.path.join(Config.TEMP_DIR, output_filename)
        
        logger.info("=" * 60)
        logger.info(f"🚀 Starting generation for: {selected_name}")
        logger.info("=" * 60)
        
        shutil.copy(spek_path, output_path)
        
        # Track timing
        generation_start = time.time()
        detection_time = st.session_state.get('last_detection_time', 0)
        processing_time = 0
        error_message = None
        
        # Enhanced progress tracking
        progress_bar = st.progress(0, text="Starting generation...")
        
        with st.status(f"⚙️ Generating {selected_name}...", expanded=True) as status:
            excel = None
            success = False
            
            try:
                # Step 1: Initialize Excel (10%)
                progress_bar.progress(10, text="🔌 Initializing Excel engine...")
                st.write("🔌 Initializing Excel engine...")
                excel = init_excel_app()
                
                # Step 2: Open workbook (20%)
                progress_bar.progress(20, text="📁 Opening workbook...")
                st.write(f"📁 Opening workbook...")
                wb = excel.Workbooks.Open(
                    os.path.abspath(output_path),
                    UpdateLinks=0,
                    ReadOnly=False,
                    IgnoreReadOnlyRecommended=True
                )
                
                excel.Calculation = ExcelConst.CALCULATION_MANUAL
                
                # Step 3: Create dump sheet (30%)
                progress_bar.progress(30, text="📝 Creating dump sheet...")
                st.write("📝 Creating dump sheet...")
                create_dump_sheet(wb)
                
                # Step 4: Inject templates (40%)
                progress_bar.progress(40, text="📑 Injecting template sheets...")
                st.write("📑 Injecting template sheets...")
                if os.path.exists(Config.TEMPLATE_FILE):
                    template_success, template_msg = inject_template_sheets(wb, Config.TEMPLATE_FILE, excel)
                    st.write(template_msg)
                else:
                    st.warning(f"⚠️ Template file not found: {Config.TEMPLATE_FILE}")
                
                # Step 5: Process cutting (70%) - longest step
                progress_bar.progress(50, text="✂️ Processing sheets...")
                st.write("✂️ Processing: cut columns & add evaluation...")
                process_start = time.time()
                process_sheet_cutting(wb, selected["col_index"], excel)
                processing_time = time.time() - process_start
                
                # Step 6: Clean sheets (80%)
                progress_bar.progress(80, text="🧹 Cleaning unused sheets...")
                st.write("🧹 Cleaning unused sheets...")
                deleted = delete_unused_sheets(wb)
                if deleted:
                    st.write(f"🗑️ Deleted: {', '.join(deleted)}")
                
                # Step 7: Finalize view (90%)
                progress_bar.progress(90, text="🎨 Finalizing view...")
                st.write("🎨 Finalizing view...")
                finalize_workbook_view(wb, excel)
                
                # Step 8: Save (95%)
                progress_bar.progress(95, text="💾 Saving file...")
                st.write("💾 Saving...")
                excel.Calculation = ExcelConst.CALCULATION_AUTOMATIC
                wb.Save()
                
                # Step 9: Verify output (98%)
                progress_bar.progress(98, text="🔍 Verifying output...")
                st.write("🔍 Verifying output quality...")
                
                output_size = os.path.getsize(output_path) if os.path.exists(output_path) else 0
                is_valid, verify_report, verify_msg = verify_output_file(output_path)
                
                if is_valid:
                    st.success(verify_msg)
                    if verify_report.get("warnings"):
                        with st.expander("⚠️ Verification Warnings"):
                            for warning in verify_report["warnings"]:
                                st.caption(f"• {warning}")
                else:
                    st.warning(verify_msg)
                    st.caption("File generated but quality verification found issues")
                
                # Complete (100%)
                progress_bar.progress(100, text="✅ Complete!")
                
                generation_duration = time.time() - generation_start
                logger.info(f"✅ Generation completed successfully: {output_filename} in {generation_duration:.2f}s")
                
                status.update(
                    label=f"✅ Complete! {output_filename} ready ({generation_duration:.1f}s)",
                    state="complete",
                    expanded=False
                )
                success = True
                
                # Record metrics if available
                if METRICS_AVAILABLE:
                    try:
                        record_generation(
                            ccode_name=selected_name,
                            success=True,
                            duration=generation_duration,
                            detection_time=detection_time,
                            processing_time=processing_time,
                            output_size=output_size
                        )
                        logger.info("📊 Metrics recorded successfully")
                    except Exception as me:
                        logger.warning(f"Failed to record metrics: {me}")
            
            except Exception as e:
                error_message = str(e)
                generation_duration = time.time() - generation_start
                
                logger.error(f"❌ Generation failed: {e}", exc_info=True)
                status.update(label="❌ Error occurred!", state="error", expanded=True)
                progress_bar.progress(0, text="❌ Failed!")
                st.error(f"Error: {e}")
                
                # Record failure metrics
                if METRICS_AVAILABLE:
                    try:
                        record_generation(
                            ccode_name=selected_name,
                            success=False,
                            duration=generation_duration,
                            error_message=error_message
                        )
                    except:
                        pass
            
            finally:
                if excel:
                    close_excel_safely(excel)
        
        if success:
            st.balloons()
            with open(output_path, "rb") as f:
                st.download_button(
                    label="📥 Download Checksheet",
                    data=f,
                    file_name=output_filename,
                    mime="application/vnd.ms-excel.sheet.macroEnabled.12"
                )
    
    # Sidebar with logs
    with st.sidebar:
        st.markdown("### 📊 System Status")
        st.write(f"🔧 openpyxl: {'✅' if OPENPYXL_AVAILABLE else '❌'}")
        st.write(f"📄 Template: {'✅' if os.path.exists(Config.TEMPLATE_FILE) else '❌'}")
        st.write(f"📈 Metrics: {'✅' if METRICS_AVAILABLE else '❌'}")
        
        if st.button("🧹 Clean Temp Files"):
            deleted = cleanup_old_temp_files(max_age_hours=0)
            st.success(f"Deleted {deleted} file(s)")
        
        # Metrics Dashboard
        if METRICS_AVAILABLE:
            st.markdown("---")
            st.markdown("### 📈 Performance Metrics")
            
            try:
                from metrics import get_metrics_collector
                collector = get_metrics_collector()
                summary = collector.get_summary()
                
                # Key metrics
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("Total", summary.get("total_generations", 0))
                    st.metric("Success Rate", f"{summary.get('success_rate', 0):.1f}%")
                with col2:
                    st.metric("Avg Time", f"{summary.get('avg_duration_seconds', 0):.1f}s")
                    fastest = summary.get('fastest_generation', {}).get('duration', 0)
                    st.metric("Fastest", f"{fastest:.1f}s" if fastest > 0 else "N/A")
                
                # Most used C-Code
                most_used = collector.get_most_used_ccode()
                if most_used:
                    ccode, count = most_used
                    st.caption(f"🎯 Most used: **{ccode}** ({count}x)")
                
                # Export button
                if st.button("📊 Export Report", key="export_metrics"):
                    report_file = f"metrics/report_{datetime.now():%Y%m%d_%H%M%S}.txt"
                    collector.export_report(report_file)
                    st.success(f"Report saved: {report_file}")
                    
            except Exception as me:
                st.caption(f"⚠️ Metrics unavailable: {me}")
        
        st.markdown("---")
        st.caption("📝 Latest log entries:")
        
        # Show recent log entries
        log_file = f"{LOG_DIR}/app_{datetime.now():%Y%m%d}.log"
        if os.path.exists(log_file):
            try:
                with open(log_file, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                    recent = lines[-10:] if len(lines) > 10 else lines
                    for line in recent:
                        st.caption(line.strip())
            except:
                pass

if __name__ == "__main__":
    main()
