"""
MIB Checksheet Generator - Standalone Integrated Version
✅ Mode 1: Model Selection (Overall Checksheet)
✅ Mode 2: Model Selection + Double Validation (Section Checksheet)
✅ Background processing dengan threading
✅ No external dependencies (standalone)
"""

import streamlit as st
import win32com.client as win32
import os
import warnings
import pythoncom
import time
import threading
import queue
import traceback
import tempfile
import shutil
import re
from datetime import datetime

# Suppress warnings
warnings.filterwarnings('ignore', message='.*ScriptRunContext.*')

# Import openpyxl if available
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ==========================================
# CONSTANTS & CONFIGURATION
# ==========================================

class ExcelConst:
    """Excel API Constants"""
    CALCULATION_MANUAL = -4135
    CALCULATION_AUTOMATIC = -4105
    SHEET_VISIBLE = -1
    PASTE_ALL = -4163
    HORIZONTAL_CENTER = -4108
    NORMAL_VIEW = 1

class Config:
    """Application Configuration"""
    # Sheet names
    SHEET_PUBLIC_MIB = "Public MIB"
    SHEET_PRIVATE_MIB = "EPSON Private MIB"
    SHEET_DUMP = "dump"
    SHEET_TEMPLATE_SUMMARY = "集計表"
    SHEET_TEMPLATE_OID = "oid"
    
    IMPORTANT_SHEETS = [SHEET_PUBLIC_MIB, SHEET_PRIVATE_MIB, SHEET_DUMP, 
                        SHEET_TEMPLATE_SUMMARY, SHEET_TEMPLATE_OID]
    
    # Column indices
    COL_START_CCODE = 11  # Column K
    COL_END_SCAN = 200
    COL_STEP = 1  # Scan all columns (changed from 5 to detect all models)
    COL_START_EVAL = 17  # Column Q
    
    # Row indices  
    ROW_HEADER = 10
    ROW_SUBHEADER = 12
    ROW_DATA_START = 13
    
    # Display settings
    DEFAULT_ZOOM = 80
    BATCH_SIZE_UNION = 500

# Evaluation column headers
EVAL_HEADERS = ["値(比較用)", "値(比較用） 加工", "取得値", "自動", "手動", 
                "判定理由", "担当者", "自/他", "手動", "判定理由"]

# Excel formulas
FORMULA_SHOURYAKU = '=IF(AND(M13="", O13="", N13<>""), "○", "")'
FORMULA_VALUE_COMPARE = '=IF(P13="","",IF(P13="←",IF(OFFSET($J13,0,MATCH(MID(M13,1,FIND("の",M13,1)-1), $L$11:$BA$11,0)+4,1,1)="","",OFFSET($L13,0,MATCH(MID(M13,1,FIND("の",M13,1)-1), $L$11:$BA$11,0)+4,1,1)),P13))'
FORMULA_VALUE_PROCESS = '=IF(COUNTIF(Q13,"*(*"),MID(Q13,FIND("(",Q13,1)+1,FIND(")",Q13,1)-FIND("(",Q13,1)-1),IF(COUNTIF(Q13,"*""*"),MID(Q13,FIND("""",Q13,1)+1,LEN(Q13)-2),IF(COUNTIF(Q13,"*：*"),RIGHT(Q13,LEN(Q13)-(FIND("：",Q13))),Q13)))'
FORMULA_VLOOKUP = '=IFERROR(IF(INDEX(dump!$D:$D,MATCH(TRIM($F13),dump!$A:$A,0))="","空文字",INDEX(dump!$D:$D,MATCH(TRIM($F13),dump!$A:$A,0))),IFERROR(IF(INDEX(dump!$D:$D,MATCH(SUBSTITUTE(SUBSTITUTE(TRIM($F13),".x",".1"),".X",".1"),dump!$A:$A,0))="","空文字",INDEX(dump!$D:$D,MATCH(SUBSTITUTE(SUBSTITUTE(TRIM($F13),".x",".1"),".X",".1"),dump!$A:$A,0))),IFERROR(IF(INDEX(dump!$D:$D,MATCH(TRIM($F13)&".1",dump!$A:$A,0))="","空文字",INDEX(dump!$D:$D,MATCH(TRIM($F13)&".1",dump!$A:$A,0))),"NA")))'
FORMULA_HANTEI = '=IF($K13<>"",IF($E13<>"","■",""),IF(AND(R13="",S13="NA"),"●", IF(EXACT(R13,S13),"●","×")))'

def _normalize_oid_for_lookup(oid_value):
    """Normalize OID for lookup by converting segment '.x' to '.1'."""
    oid = str(oid_value).strip() if oid_value is not None else ""
    if not oid:
        return ""
    # Replace only OID segment '.x' or '.X' before '.' or end-of-string.
    return re.sub(r'\.[xX](?=\.|$)', '.1', oid)

def _load_auto_maru_oids_by_section():
    """Load OID list per section from oids_auto_maru.txt.
    Supported section examples: [PUBLIC_MIB], [EPSON_PRIVATE_MIB]."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    txt_path = os.path.join(base_dir, "oids_auto_maru.txt")
    section_to_oids = {
        "PUBLIC_MIB": set(),
        "EPSON_PRIVATE_MIB": set(),
    }
    current_section = None

    try:
        with open(txt_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue

                if line.startswith("[") and line.endswith("]"):
                    current_section = line[1:-1].strip().upper()
                    if current_section not in section_to_oids:
                        section_to_oids[current_section] = set()
                    continue

                if current_section:
                    section_to_oids[current_section].add(line)
    except FileNotFoundError:
        pass

    frozen = {k: frozenset(v) for k, v in section_to_oids.items()}
    return frozen

AUTO_MARU_OIDS_BY_SECTION = _load_auto_maru_oids_by_section()

# ==========================================
# UTILITY FUNCTIONS
# ==========================================

def extract_numeric_value(value_str):
    """
    Extract numeric value from various string formats.
    Examples:
        "初期値：0" → "0"
        "impression(3)" → "3"
        "100" → "100"
        "default:5" → "5"
        "0x0A" → "10" (converts hex to decimal)
        "空文字" → "" (empty for special values)
    """
    if not value_str or value_str is None:
        return ""
    
    value_str = str(value_str).strip()
    
    # Special cases - return empty for non-numeric strings
    if value_str.lower() in ["", "空文字", "na", "n/a", "-"]:
        return ""
    
    # Try to extract number from common patterns
    # Pattern 1: 初期値：123 or default:123 or value:123
    match = re.search(r'[：:]\s*(-?\d+(?:\.\d+)?)', value_str)
    if match:
        return match.group(1)
    
    # Pattern 2: impression(123) or function(123)
    match = re.search(r'\((-?\d+(?:\.\d+)?)\)', value_str)
    if match:
        return match.group(1)
    
    # Pattern 3: Hex value like 0x0A or 0X0A
    match = re.search(r'0[xX]([0-9A-Fa-f]+)', value_str)
    if match:
        try:
            return str(int(match.group(1), 16))
        except:
            pass
    
    # Pattern 4: Just extract any number (including negative and decimal)
    match = re.search(r'(-?\d+(?:\.\d+)?)', value_str)
    if match:
        return match.group(1)
    
    # If no number found, return original value
    return value_str

def detect_sheet_names(checksheet_filename):
    """
    Detect sheet names based on checksheet filename.
    
    Returns: (checksheet_sheet_name, mib_spek_sheet_name)
    
    Rules:
    - epPrt → Checksheet: "Printer Private", MIB: "EPSON Private MIB"
    - epScn → Checksheet: "Scanner Private", MIB: "EPSON Private MIB"
    - epHr → Checksheet: "HR Private", MIB: "EPSON Private MIB"
    - public → Checksheet: "Printer Public", MIB: "Public MIB"
    - default → Checksheet: "Printer Private", MIB: "EPSON Private MIB"
    """
    filename_lower = checksheet_filename.lower()
    
    if 'public' in filename_lower:
        return "Printer Public", "Public MIB"
    elif 'epscn' in filename_lower:
        return "Scanner Private", "EPSON Private MIB"
    elif 'ephr' in filename_lower:
        return "HR Private", "EPSON Private MIB"
    elif 'epprt' in filename_lower:
        return "Printer Private", "EPSON Private MIB"
    elif 'ep' in filename_lower:
        # Generic 'ep' fallback - use EPSON Private MIB
        return "Printer Private", "EPSON Private MIB"
    else:
        # Default fallback
        return "Printer Private", "EPSON Private MIB"

# ==========================================
# EXCEL CORE OPERATIONS
# ==========================================

def init_excel_app():
    """Initialize Excel application in background mode."""
    pythoncom.CoInitialize()
    excel = create_excel_app()
    
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    excel.EnableEvents = False
    return excel

def create_excel_app():
    """Create Excel COM app without relying on EnsureDispatch (gen_py can be corrupt)."""
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        return excel
    except Exception:
        pass

    try:
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        return excel
    except Exception as first_error:
        # Last resort: clear generated COM cache and retry.
        try:
            import win32com
            gen_path = getattr(win32com, "__gen_path__", "")
            if gen_path and os.path.isdir(gen_path):
                shutil.rmtree(gen_path, ignore_errors=True)

            try:
                win32.gencache.is_readonly = False
                win32.gencache.Rebuild()
            except Exception:
                pass

            excel = win32.Dispatch("Excel.Application")
            excel.Visible = False
            return excel
        except Exception as final_error:
            raise RuntimeError(
                f"Failed to initialize Excel COM. First error: {first_error}; after cache reset: {final_error}"
            )

def close_excel_safely(excel, wb=None):
    """Close Excel application and cleanup COM."""
    try:
        while excel.Workbooks.Count > 0:
            excel.Workbooks(1).Close(SaveChanges=False)
    except:
        pass
    
    try:
        excel.Visible = False  # Ensure hidden before quit
        excel.ScreenUpdating = True
        excel.EnableEvents = True
        excel.DisplayAlerts = True
        excel.Quit()
    except:
        pass
    
    try:
        pythoncom.CoUninitialize()
    except:
        pass

def find_last_row(sheet):
    """Find the last used row in a sheet."""
    try:
        if sheet.UsedRange.Rows.Count > 0:
            last_row = sheet.UsedRange.Rows.Count + sheet.UsedRange.Row - 1
            if last_row > 10:
                return last_row
        
        last_cell = sheet.Cells.Find(What="*", SearchOrder=1, SearchDirection=2)
        if last_cell and last_cell.Row > 10:
            return last_cell.Row
    except:
        pass
    
    return 500

# ==========================================
# MODEL DETECTION (MODE 1)
# ==========================================

def read_ccodes_openpyxl(file_path):
    """Fast Model detection using openpyxl (no Excel needed)."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb[Config.SHEET_PUBLIC_MIB]
    ccodes = []
    empty_count = 0
    
    for col_idx in range(Config.COL_START_CCODE, Config.COL_END_SCAN, Config.COL_STEP):
        value = sheet.cell(row=Config.ROW_HEADER, column=col_idx).value
        if not value or str(value).strip() == "":
            empty_count += 1
            if empty_count >= 20:  # Stop after 20 consecutive empty columns
                break
            continue
        
        empty_count = 0  # Reset counter when non-empty column found
        ccodes.append({"name": str(value).strip(), "col_index": col_idx})
    
    wb.close()
    return ccodes

# ==========================================
# SHEET OPERATIONS (MODE 1)
# ==========================================

def create_dump_sheet(wb):
    """Create empty dump sheet for VLOOKUP."""
    try:
        target_sheet = wb.Sheets(Config.SHEET_PRIVATE_MIB)
        new_sheet = wb.Sheets.Add(After=target_sheet)
    except:
        new_sheet = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
    
    new_sheet.Name = Config.SHEET_DUMP

def insert_shouryaku_column(sheet, last_row):
    """Insert 省略 column."""
    sheet.Columns(Config.COL_START_CCODE).Insert()
    
    header_cell = sheet.Cells(Config.ROW_HEADER, Config.COL_START_CCODE)
    header_cell.Value = "省略"
    header_cell.Font.Bold = False
    header_cell.Font.Size = 9
    
    merge_area = sheet.Range(
        sheet.Cells(9, Config.COL_START_CCODE),
        sheet.Cells(Config.ROW_SUBHEADER, Config.COL_START_CCODE)
    )
    merge_area.Merge()
    
    format_area = sheet.Range(
        sheet.Cells(9, Config.COL_START_CCODE),
        sheet.Cells(last_row, Config.COL_START_CCODE)
    )
    format_area.Interior.ColorIndex = 2
    format_area.Borders.LineStyle = 1
    format_area.Borders.Weight = 2
    format_area.HorizontalAlignment = ExcelConst.HORIZONTAL_CENTER

def delete_unused_columns(sheet, target_col):
    """Delete columns outside the target Model."""
    new_target = target_col + 1
    delete_start = new_target + 5
    
    sheet.Range(
        sheet.Columns(delete_start),
        sheet.Columns(Config.COL_END_SCAN)
    ).Delete()
    
    if new_target > 12:
        sheet.Range(
            sheet.Columns(12),
            sheet.Columns(new_target - 1)
        ).Delete()

def add_evaluation_columns(sheet, last_row, excel_app):
    """Add evaluation columns with formulas."""
    start_col = Config.COL_START_EVAL
    
    original_calc = excel_app.Calculation
    excel_app.Calculation = ExcelConst.CALCULATION_MANUAL
    
    try:
        clear_area = sheet.Range(
            sheet.Cells(9, start_col),
            sheet.Cells(min(last_row + 10, 5000), start_col + 9)
        )
        clear_area.ClearFormats()
        
        header_range = sheet.Range(
            sheet.Cells(Config.ROW_SUBHEADER, start_col),
            sheet.Cells(Config.ROW_SUBHEADER, start_col + len(EVAL_HEADERS) - 1)
        )
        header_range.Value = tuple(EVAL_HEADERS)
        header_range.Font.Bold = False
        
        table_area = sheet.Range(
            sheet.Cells(9, start_col),
            sheet.Cells(last_row, start_col + 9)
        )
        table_area.Interior.ColorIndex = 2
        table_area.Borders.LineStyle = 1
        table_area.Borders.Weight = 2
        table_area.ColumnWidth = 12
        table_area.WrapText = True
        table_area.Font.Size = 9
        
        if last_row >= Config.ROW_DATA_START:
            _apply_evaluation_formulas(sheet, last_row)

            # Read col B(2)–M(13) in ONE COM call.
            # B=idx0, C=idx1, D=idx2, F=idx4, J=idx8, M=idx11
            raw = sheet.Range(
                sheet.Cells(Config.ROW_DATA_START, 2),
                sheet.Cells(last_row, 13)
            ).Value
            if raw is None:
                raw = []
            elif not isinstance(raw, tuple) or (len(raw) > 0 and not isinstance(raw[0], tuple)):
                raw = [(raw,)]

            _clear_nw_rows(sheet, last_row, raw)
            _apply_auto_maru_u_by_sheet_oid_list(sheet, last_row, raw)
    finally:
        excel_app.Calculation = original_calc

def _apply_evaluation_formulas(sheet, last_row):
    """Apply evaluation formulas to ranges."""
    area_k = sheet.Range(
        sheet.Cells(Config.ROW_DATA_START, 11),
        sheet.Cells(last_row, 11)
    )
    area_k.Formula = FORMULA_SHOURYAKU
    
    areas = [
        (17, FORMULA_VALUE_COMPARE),
        (18, FORMULA_VALUE_PROCESS),
        (19, FORMULA_VLOOKUP),
        (20, FORMULA_HANTEI)
    ]
    
    for col_idx, formula in areas:
        area = sheet.Range(
            sheet.Cells(Config.ROW_DATA_START, col_idx),
            sheet.Cells(last_row, col_idx)
        )
        area.Formula = formula

def _clear_nw_rows(sheet, last_row, data_b_to_j):
    """Clear Q:Z when row is NW, B/C/D contain values, or F OID has no dot.
    data_b_to_j: pre-read tuple-of-tuples for cols B(idx0)–M(idx11)."""
    if not data_b_to_j:
        return

    ranges_to_clear = []
    total_rows = last_row - Config.ROW_DATA_START + 1
    for i in range(total_rows):
        row = data_b_to_j[i] if i < len(data_b_to_j) else None
        if not row:
            continue

        # col J = index 8
        value_j = str(row[8]) if row[8] is not None else ""
        # col B-D = indices 0-2
        has_value_bcd = any(row[k] not in (None, "") for k in range(3))
        # col F = index 4
        value_f = str(row[4]).strip() if row[4] is not None else ""
        is_no_dot_oid = bool(value_f and "." not in value_f)

        if "NW" in value_j or has_value_bcd or is_no_dot_oid:
            row_num = Config.ROW_DATA_START + i
            ranges_to_clear.append(f"Q{row_num}:Z{row_num}")
    
    if ranges_to_clear:
        for batch_start in range(0, len(ranges_to_clear), Config.BATCH_SIZE_UNION):
            batch = ranges_to_clear[batch_start:batch_start + Config.BATCH_SIZE_UNION]
            try:
                combined = sheet.Range(batch[0])
                for addr in batch[1:]:
                    try:
                        combined = sheet.Application.Union(combined, sheet.Range(addr))
                    except:
                        pass
                combined.ClearContents()
            except:
                pass

def _apply_auto_maru_u_by_sheet_oid_list(sheet, last_row, data_b_to_j):
    """Fill column U with ● for OIDs listed in oids_auto_maru.txt based on current sheet."""
    if last_row < Config.ROW_DATA_START or not data_b_to_j:
        return

    sheet_name = str(sheet.Name).strip().lower()
    if sheet_name == Config.SHEET_PUBLIC_MIB.lower():
        target_oids = AUTO_MARU_OIDS_BY_SECTION.get("PUBLIC_MIB", frozenset())
    elif sheet_name == Config.SHEET_PRIVATE_MIB.lower():
        target_oids = AUTO_MARU_OIDS_BY_SECTION.get("EPSON_PRIVATE_MIB", frozenset())
    else:
        target_oids = frozenset()

    normalized_target_oids = {_normalize_oid_for_lookup(oid) for oid in target_oids}

    total_rows = last_row - Config.ROW_DATA_START + 1
    rows_for_u = []

    for i in range(total_rows):
        row = data_b_to_j[i] if i < len(data_b_to_j) else None
        if not row:
            continue

        # col F = index 4 in B..J range
        oid = _normalize_oid_for_lookup(row[4])

        if oid in normalized_target_oids:
            rows_for_u.append(Config.ROW_DATA_START + i)

    # Set value directly to ● in column U (21)
    for batch_start in range(0, len(rows_for_u), Config.BATCH_SIZE_UNION):
        batch = rows_for_u[batch_start:batch_start + Config.BATCH_SIZE_UNION]
        if not batch:
            continue
        try:
            if len(batch) == 1:
                sheet.Cells(batch[0], 21).Value = "●"
            else:
                combined = sheet.Cells(batch[0], 21)
                for row_num in batch[1:]:
                    combined = sheet.Application.Union(combined, sheet.Cells(row_num, 21))
                combined.Value = "●"
        except:
            pass

    # Sync Y from U: every row where U is ●, set Y to ●
    u_data = sheet.Range(
        sheet.Cells(Config.ROW_DATA_START, 21),
        sheet.Cells(last_row, 21)
    ).Value

    if u_data is None:
        u_data = []
    elif not isinstance(u_data, tuple) or (len(u_data) > 0 and not isinstance(u_data[0], tuple)):
        u_data = [(u_data,)]

    rows_for_y = []
    for i in range(total_rows):
        row_u = u_data[i] if i < len(u_data) else None
        value_u = str(row_u[0]).strip() if row_u and row_u[0] is not None else ""
        if value_u == "●":
            rows_for_y.append(Config.ROW_DATA_START + i)

    for batch_start in range(0, len(rows_for_y), Config.BATCH_SIZE_UNION):
        batch = rows_for_y[batch_start:batch_start + Config.BATCH_SIZE_UNION]
        if not batch:
            continue
        try:
            if len(batch) == 1:
                sheet.Cells(batch[0], 25).Value = "●"
            else:
                combined = sheet.Cells(batch[0], 25)
                for row_num in batch[1:]:
                    combined = sheet.Application.Union(combined, sheet.Cells(row_num, 25))
                combined.Value = "●"
        except:
            pass

def delete_unused_sheets(wb):
    """Delete sheets not in important list."""
    target_lower = [s.lower() for s in Config.IMPORTANT_SHEETS]
    deleted = []
    
    for sheet in wb.Sheets:
        if sheet.Name.lower() not in target_lower:
            try:
                sheet.Delete()
                deleted.append(sheet.Name)
            except:
                pass
    
    return deleted

def unhide_all_cells(sheet):
    """Unhide all hidden columns and rows in sheet."""
    try:
        sheet.Columns.Hidden = False
        sheet.Rows.Hidden = False
    except:
        pass

def inject_template_sheets(wb_target, template_path, excel):
    """Inject template sheets from Template Sheet.xlsm."""
    if not os.path.exists(template_path):
        return False, f"❌ Template not found: {template_path}"
    
    wb_template = None
    try:
        wb_template = excel.Workbooks.Open(os.path.abspath(template_path), ReadOnly=True)
        available = [s.Name for s in wb_template.Sheets]
        target_sheets = [Config.SHEET_TEMPLATE_SUMMARY, Config.SHEET_TEMPLATE_OID]
        
        copied = []
        
        for target_name in target_sheets:
            source_name = None
            for avail in available:
                if avail.lower() == target_name.lower():
                    source_name = avail
                    break
            
            if not source_name:
                continue
            
            try:
                source_sheet = wb_template.Sheets(source_name)
                source_sheet.Copy(Before=wb_target.Sheets(1))
                
                new_sheet = wb_target.Sheets(1)
                if new_sheet.Name != target_name:
                    try:
                        new_sheet.Name = target_name
                    except:
                        pass
                
                try:
                    new_sheet.Move(After=wb_target.Sheets(wb_target.Sheets.Count))
                except:
                    pass
                
                copied.append(target_name)
                
            except:
                pass
        
        wb_template.Close(SaveChanges=False)
        
        if copied:
            return True, f"✅ Copied {len(copied)} sheets"
        else:
            return False, "❌ Failed to copy sheets"
    
    except Exception as e:
        if wb_template:
            try:
                wb_template.Close(SaveChanges=False)
            except:
                pass
        return False, f"❌ Error: {str(e)}"

def process_sheet_cutting(wb, target_col, excel):
    """Main processing: cut columns and add evaluation."""
    target_sheets = [Config.SHEET_PUBLIC_MIB, Config.SHEET_PRIVATE_MIB]
    
    original_screen = excel.ScreenUpdating
    original_calc = excel.Calculation
    
    try:
        excel.ScreenUpdating = False
        excel.Calculation = ExcelConst.CALCULATION_MANUAL
        
        for sheet_name in target_sheets:
            sheet = wb.Sheets(sheet_name)
            last_row = find_last_row(sheet)
            
            insert_shouryaku_column(sheet, last_row)
            delete_unused_columns(sheet, target_col)
            add_evaluation_columns(sheet, last_row, excel)
            unhide_all_cells(sheet)
            
            sheet.Activate()
            sheet.PageSetup.PrintArea = ""
    finally:
        excel.ScreenUpdating = original_screen
        excel.Calculation = original_calc

# ==========================================
# MODEL DETECTION (MODE 2)
# ==========================================

def detect_models_from_spek(file_spek, checksheet_filename=""):
    """Detect available models from Spek file."""
    pythoncom.CoInitialize()
    excel = create_excel_app()
    
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    excel.EnableEvents = False
    
    models = []
    
    # Detect which MIB sheet to use based on checksheet filename
    _, mib_sheet_name = detect_sheet_names(checksheet_filename)
    
    try:
        wb = excel.Workbooks.Open(os.path.abspath(file_spek), ReadOnly=True)
        ws = wb.Worksheets(mib_sheet_name)
        
        COL_START = 11
        COL_END = 200
        COL_STEP = 1  # Scan all columns
        ROW_HEADER = 10
        
        empty_count = 0
        for col_idx in range(COL_START, COL_END, COL_STEP):
            header_value = ws.Cells(ROW_HEADER, col_idx).Value
            if not header_value or str(header_value).strip() == "":
                empty_count += 1
                if empty_count >= 20:  # Stop after 20 consecutive empty columns
                    break
                continue
            
            empty_count = 0  # Reset counter
            model_name = str(header_value).strip()
            models.append({
                "name": model_name,
                "col_index": col_idx,
                "col_atchi": col_idx + 4
            })
        
        wb.Close(False)
        
    except Exception as e:
        pass
    finally:
        try:
            if 'excel' in locals():
                try:
                    # Close all open workbooks
                    while excel.Workbooks.Count > 0:
                        excel.Workbooks(1).Close(SaveChanges=False)
                except:
                    pass
                
                try:
                    excel.Visible = False  # Ensure hidden
                    excel.ScreenUpdating = True
                    excel.Quit()
                except:
                    pass
        except:
            pass
        
        try:
            pythoncom.CoUninitialize()
        except:
            pass
    
    return models

# ==========================================
# BACKGROUND WORKERS
# ==========================================

