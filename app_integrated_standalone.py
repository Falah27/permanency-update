"""
MIB Checksheet Generator - Standalone Integrated Version
✅ Mode 1: Model Selection (Overall Checksheet)
✅ Mode 2: Model Selection + Double Validation (Section Checksheet)
✅ Background processing dengan threading
✅ No external dependencies (standalone)
"""

import streamlit as st
import win32com.client as win32
import os
import sys
import warnings
import pythoncom
import time
import threading
import queue
import traceback
import tempfile
import shutil
import re
from datetime import datetime

# Suppress warnings
warnings.filterwarnings('ignore', message='.*ScriptRunContext.*')

# Import openpyxl if available
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ==========================================
# CONSTANTS & CONFIGURATION
# ==========================================

class ExcelConst:
    """Excel API Constants"""
    CALCULATION_MANUAL = -4135
    CALCULATION_AUTOMATIC = -4105
    SHEET_VISIBLE = -1
    PASTE_ALL = -4163
    HORIZONTAL_CENTER = -4108
    NORMAL_VIEW = 1

class Config:
    """Application Configuration"""
    # Sheet names
    SHEET_PUBLIC_MIB = "Public MIB"
    SHEET_PRIVATE_MIB = "EPSON Private MIB"
    SHEET_DUMP = "dump"
    SHEET_TEMPLATE_SUMMARY = "集計表"
    SHEET_TEMPLATE_OID = "oid"
    
    IMPORTANT_SHEETS = [SHEET_PUBLIC_MIB, SHEET_PRIVATE_MIB, SHEET_DUMP, 
                        SHEET_TEMPLATE_SUMMARY, SHEET_TEMPLATE_OID]
    
    # Column indices
    COL_START_CCODE = 11  # Column K
    COL_END_SCAN = 200
    COL_STEP = 1  # Scan all columns (changed from 5 to detect all models)
    COL_START_EVAL = 17  # Column Q
    
    # Row indices  
    ROW_HEADER = 10
    ROW_SUBHEADER = 12
    ROW_DATA_START = 13
    
    # Display settings
    DEFAULT_ZOOM = 80
    BATCH_SIZE_UNION = 500

# Evaluation column headers
EVAL_HEADERS = ["値(比較用)", "値(比較用） 加工", "取得値", "自動", "手動", 
                "判定理由", "担当者", "自/他", "手動", "判定理由"]

# Excel formulas
FORMULA_SHOURYAKU = '=IF(AND(M13="", O13="", N13<>""), "○", "")'
FORMULA_VALUE_COMPARE = '=IF(P13="","",IF(P13="←",IF(OFFSET($J13,0,MATCH(MID(M13,1,FIND("の",M13,1)-1), $L$11:$BA$11,0)+4,1,1)="","",OFFSET($L13,0,MATCH(MID(M13,1,FIND("の",M13,1)-1), $L$11:$BA$11,0)+4,1,1)),P13))'
FORMULA_VALUE_PROCESS = '=IF(COUNTIF(Q13,"*(*"),MID(Q13,FIND("(",Q13,1)+1,FIND(")",Q13,1)-FIND("(",Q13,1)-1),IF(COUNTIF(Q13,"*""*"),MID(Q13,FIND("""",Q13,1)+1,LEN(Q13)-2),IF(COUNTIF(Q13,"*：*"),RIGHT(Q13,LEN(Q13)-(FIND("：",Q13))),Q13)))'
FORMULA_VLOOKUP = '=IFERROR(IF(VLOOKUP("*"&TRIM($F13)&"*",dump!$A:$D,4,FALSE)="","空文字",VLOOKUP("*"&TRIM($F13)&"*",dump!$A:$D,4,FALSE)),"NA")'
FORMULA_HANTEI = '=IF($K13<>"",IF($E13<>"","■",""),IF(AND(R13="",S13="NA"),"●", IF(EXACT(R13,S13),"●","×")))'

# ==========================================
# UTILITY FUNCTIONS
# ==========================================

def extract_numeric_value(value_str):
    """
    Extract numeric value from various string formats.
    Examples:
        "初期値：0" → "0"
        "impression(3)" → "3"
        "100" → "100"
        "default:5" → "5"
        "0x0A" → "10" (converts hex to decimal)
        "空文字" → "" (empty for special values)
    """
    if not value_str or value_str is None:
        return ""
    
    value_str = str(value_str).strip()
    
    # Special cases - return empty for non-numeric strings
    if value_str.lower() in ["", "空文字", "na", "n/a", "-"]:
        return ""
    
    # Try to extract number from common patterns
    # Pattern 1: 初期値：123 or default:123 or value:123
    match = re.search(r'[：:]\s*(-?\d+(?:\.\d+)?)', value_str)
    if match:
        return match.group(1)
    
    # Pattern 2: impression(123) or function(123)
    match = re.search(r'\((-?\d+(?:\.\d+)?)\)', value_str)
    if match:
        return match.group(1)
    
    # Pattern 3: Hex value like 0x0A or 0X0A
    match = re.search(r'0[xX]([0-9A-Fa-f]+)', value_str)
    if match:
        try:
            return str(int(match.group(1), 16))
        except:
            pass
    
    # Pattern 4: Just extract any number (including negative and decimal)
    match = re.search(r'(-?\d+(?:\.\d+)?)', value_str)
    if match:
        return match.group(1)
    
    # If no number found, return original value
    return value_str

def detect_sheet_names(checksheet_filename):
    """
    Detect sheet names based on checksheet filename.
    
    Returns: (checksheet_sheet_name, mib_spek_sheet_name)
    
    Rules:
    - epPrt → Checksheet: "Printer Private", MIB: "EPSON Private MIB"
    - epScn → Checksheet: "Scanner Private", MIB: "EPSON Private MIB"
    - epHr → Checksheet: "HR Private", MIB: "EPSON Private MIB"
    - public → Checksheet: "Printer Public", MIB: "Public MIB"
    - default → Checksheet: "Printer Private", MIB: "EPSON Private MIB"
    """
    filename_lower = checksheet_filename.lower()
    
    if 'public' in filename_lower:
        return "Printer Public", "Public MIB"
    elif 'epscn' in filename_lower:
        return "Scanner Private", "EPSON Private MIB"
    elif 'ephr' in filename_lower:
        return "HR Private", "EPSON Private MIB"
    elif 'epprt' in filename_lower:
        return "Printer Private", "EPSON Private MIB"
    elif 'ep' in filename_lower:
        # Generic 'ep' fallback - use EPSON Private MIB
        return "Printer Private", "EPSON Private MIB"
    else:
        # Default fallback
        return "Printer Private", "EPSON Private MIB"

# ==========================================
# EXCEL CORE OPERATIONS
# ==========================================

def init_excel_app():
    """Initialize Excel application in background mode."""
    pythoncom.CoInitialize()
    try:
        excel = win32.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = False  # Set immediately
    except:
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False  # Set immediately
    
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    excel.EnableEvents = False
    return excel

def close_excel_safely(excel, wb=None):
    """Close Excel application and cleanup COM."""
    try:
        workbook_count = excel.Workbooks.Count
        while excel.Workbooks.Count > 0:
            excel.Workbooks(1).Close(SaveChanges=False)
    except:
        pass
    
    try:
        excel.Visible = False  # Ensure hidden before quit
        excel.ScreenUpdating = True
        excel.EnableEvents = True
        excel.DisplayAlerts = True
        excel.Quit()
    except:
        pass
    
    try:
        pythoncom.CoUninitialize()
    except:
        pass

def find_last_row(sheet):
    """Find the last used row in a sheet."""
    try:
        if sheet.UsedRange.Rows.Count > 0:
            last_row = sheet.UsedRange.Rows.Count + sheet.UsedRange.Row - 1
            if last_row > 10:
                return last_row
        
        last_cell = sheet.Cells.Find(What="*", SearchOrder=1, SearchDirection=2)
        if last_cell and last_cell.Row > 10:
            return last_cell.Row
    except:
        pass
    
    return 500

# ==========================================
# MODEL DETECTION (MODE 1)
# ==========================================

def read_ccodes_openpyxl(file_path):
    """Fast Model detection using openpyxl (no Excel needed)."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb[Config.SHEET_PUBLIC_MIB]
    ccodes = []
    empty_count = 0
    
    for col_idx in range(Config.COL_START_CCODE, Config.COL_END_SCAN, Config.COL_STEP):
        value = sheet.cell(row=Config.ROW_HEADER, column=col_idx).value
        if not value or str(value).strip() == "":
            empty_count += 1
            if empty_count >= 20:  # Stop after 20 consecutive empty columns
                break
            continue
        
        empty_count = 0  # Reset counter when non-empty column found
        ccodes.append({"name": str(value).strip(), "col_index": col_idx})
    
    wb.close()
    return ccodes

# ==========================================
# SHEET OPERATIONS (MODE 1)
# ==========================================

def create_dump_sheet(wb):
    """Create empty dump sheet for VLOOKUP."""
    try:
        target_sheet = wb.Sheets(Config.SHEET_PRIVATE_MIB)
        new_sheet = wb.Sheets.Add(After=target_sheet)
    except:
        new_sheet = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
    
    new_sheet.Name = Config.SHEET_DUMP

def insert_shouryaku_column(sheet, last_row):
    """Insert 省略 column."""
    sheet.Columns(Config.COL_START_CCODE).Insert()
    
    header_cell = sheet.Cells(Config.ROW_HEADER, Config.COL_START_CCODE)
    header_cell.Value = "省略"
    header_cell.Font.Bold = False
    header_cell.Font.Size = 9
    
    merge_area = sheet.Range(
        sheet.Cells(9, Config.COL_START_CCODE),
        sheet.Cells(Config.ROW_SUBHEADER, Config.COL_START_CCODE)
    )
    merge_area.Merge()
    
    format_area = sheet.Range(
        sheet.Cells(9, Config.COL_START_CCODE),
        sheet.Cells(last_row, Config.COL_START_CCODE)
    )
    format_area.Interior.ColorIndex = 2
    format_area.Borders.LineStyle = 1
    format_area.Borders.Weight = 2
    format_area.HorizontalAlignment = ExcelConst.HORIZONTAL_CENTER

def delete_unused_columns(sheet, target_col):
    """Delete columns outside the target Model."""
    new_target = target_col + 1
    delete_start = new_target + 5
    
    sheet.Range(
        sheet.Columns(delete_start),
        sheet.Columns(Config.COL_END_SCAN)
    ).Delete()
    
    if new_target > 12:
        sheet.Range(
            sheet.Columns(12),
            sheet.Columns(new_target - 1)
        ).Delete()

def add_evaluation_columns(sheet, last_row, excel_app):
    """Add evaluation columns with formulas."""
    start_col = Config.COL_START_EVAL
    
    original_calc = excel_app.Calculation
    excel_app.Calculation = ExcelConst.CALCULATION_MANUAL
    
    try:
        clear_area = sheet.Range(
            sheet.Cells(1, start_col),
            sheet.Cells(min(last_row + 10, 5000), start_col + 9)
        )
        clear_area.ClearFormats()
        
        header_range = sheet.Range(
            sheet.Cells(Config.ROW_SUBHEADER, start_col),
            sheet.Cells(Config.ROW_SUBHEADER, start_col + len(EVAL_HEADERS) - 1)
        )
        header_range.Value = tuple(EVAL_HEADERS)
        header_range.Font.Bold = False
        
        table_area = sheet.Range(
            sheet.Cells(9, start_col),
            sheet.Cells(last_row, start_col + 9)
        )
        table_area.Interior.ColorIndex = 2
        table_area.Borders.LineStyle = 1
        table_area.Borders.Weight = 2
        table_area.ColumnWidth = 12
        table_area.WrapText = True
        table_area.Font.Size = 9
        
        if last_row >= Config.ROW_DATA_START:
            _apply_evaluation_formulas(sheet, last_row)
            _clear_nw_rows(sheet, last_row)
    finally:
        excel_app.Calculation = original_calc

def _apply_evaluation_formulas(sheet, last_row):
    """Apply evaluation formulas to ranges."""
    area_k = sheet.Range(
        sheet.Cells(Config.ROW_DATA_START, 11),
        sheet.Cells(last_row, 11)
    )
    area_k.Formula = FORMULA_SHOURYAKU
    
    areas = [
        (17, FORMULA_VALUE_COMPARE),
        (18, FORMULA_VALUE_PROCESS),
        (19, FORMULA_VLOOKUP),
        (20, FORMULA_HANTEI)
    ]
    
    for col_idx, formula in areas:
        area = sheet.Range(
            sheet.Cells(Config.ROW_DATA_START, col_idx),
            sheet.Cells(last_row, col_idx)
        )
        area.Formula = formula

def _clear_nw_rows(sheet, last_row):
    """Clear evaluation columns for NW rows."""
    data_j = sheet.Range(
        sheet.Cells(Config.ROW_DATA_START, 10),
        sheet.Cells(last_row, 10)
    ).Value
    
    if not data_j:
        return
    
    ranges_to_clear = []
    for i, row_data in enumerate(data_j):
        value = str(row_data[0]) if row_data[0] else ""
        if "NW" in value:
            row_num = Config.ROW_DATA_START + i
            ranges_to_clear.append(f"Q{row_num}:Z{row_num}")
    
    if ranges_to_clear:
        for batch_start in range(0, len(ranges_to_clear), Config.BATCH_SIZE_UNION):
            batch = ranges_to_clear[batch_start:batch_start + Config.BATCH_SIZE_UNION]
            try:
                combined = sheet.Range(batch[0])
                for addr in batch[1:]:
                    try:
                        combined = sheet.Application.Union(combined, sheet.Range(addr))
                    except:
                        pass
                combined.ClearContents()
            except:
                pass

def delete_unused_sheets(wb):
    """Delete sheets not in important list."""
    target_lower = [s.lower() for s in Config.IMPORTANT_SHEETS]
    deleted = []
    
    for sheet in wb.Sheets:
        if sheet.Name.lower() not in target_lower:
            try:
                sheet.Delete()
                deleted.append(sheet.Name)
            except:
                pass
    
    return deleted

def unhide_all_cells(sheet):
    """Unhide all hidden columns and rows in sheet."""
    try:
        sheet.Columns.Hidden = False
        sheet.Rows.Hidden = False
    except:
        pass

def inject_template_sheets(wb_target, template_path, excel):
    """Inject template sheets from Template Sheet.xlsm."""
    if not os.path.exists(template_path):
        return False, f"❌ Template not found: {template_path}"
    
    wb_template = None
    try:
        wb_template = excel.Workbooks.Open(os.path.abspath(template_path), ReadOnly=True)
        available = [s.Name for s in wb_template.Sheets]
        target_sheets = [Config.SHEET_TEMPLATE_SUMMARY, Config.SHEET_TEMPLATE_OID]
        
        copied = []
        
        for target_name in target_sheets:
            source_name = None
            for avail in available:
                if avail.lower() == target_name.lower():
                    source_name = avail
                    break
            
            if not source_name:
                continue
            
            try:
                source_sheet = wb_template.Sheets(source_name)
                source_sheet.Copy(Before=wb_target.Sheets(1))
                
                new_sheet = wb_target.Sheets(1)
                if new_sheet.Name != target_name:
                    try:
                        new_sheet.Name = target_name
                    except:
                        pass
                
                try:
                    new_sheet.Move(After=wb_target.Sheets(wb_target.Sheets.Count))
                except:
                    pass
                
                copied.append(target_name)
                
            except:
                pass
        
        wb_template.Close(SaveChanges=False)
        
        if copied:
            return True, f"✅ Copied {len(copied)} sheets"
        else:
            return False, "❌ Failed to copy sheets"
    
    except Exception as e:
        if wb_template:
            try:
                wb_template.Close(SaveChanges=False)
            except:
                pass
        return False, f"❌ Error: {str(e)}"

def process_sheet_cutting(wb, target_col, excel):
    """Main processing: cut columns and add evaluation."""
    target_sheets = [Config.SHEET_PUBLIC_MIB, Config.SHEET_PRIVATE_MIB]
    
    original_screen = excel.ScreenUpdating
    original_calc = excel.Calculation
    
    try:
        excel.ScreenUpdating = False
        excel.Calculation = ExcelConst.CALCULATION_MANUAL
        
        for sheet_name in target_sheets:
            sheet = wb.Sheets(sheet_name)
            last_row = find_last_row(sheet)
            
            insert_shouryaku_column(sheet, last_row)
            delete_unused_columns(sheet, target_col)
            add_evaluation_columns(sheet, last_row, excel)
            unhide_all_cells(sheet)
            
            sheet.Activate()
            sheet.PageSetup.PrintArea = ""
    finally:
        excel.ScreenUpdating = original_screen
        excel.Calculation = original_calc

# ==========================================
# MODEL DETECTION (MODE 2)
# ==========================================

def detect_models_from_spek(file_spek, checksheet_filename=""):
    """Detect available models from Spek file."""
    pythoncom.CoInitialize()
    
    try:
        excel = win32.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = False  # Set immediately
    except:
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False  # Set immediately
    
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    excel.EnableEvents = False
    
    models = []
    
    # Detect which MIB sheet to use based on checksheet filename
    _, mib_sheet_name = detect_sheet_names(checksheet_filename)
    
    try:
        wb = excel.Workbooks.Open(os.path.abspath(file_spek), ReadOnly=True)
        ws = wb.Worksheets(mib_sheet_name)
        
        COL_START = 11
        COL_END = 200
        COL_STEP = 1  # Scan all columns
        ROW_HEADER = 10
        
        empty_count = 0
        for col_idx in range(COL_START, COL_END, COL_STEP):
            header_value = ws.Cells(ROW_HEADER, col_idx).Value
            if not header_value or str(header_value).strip() == "":
                empty_count += 1
                if empty_count >= 20:  # Stop after 20 consecutive empty columns
                    break
                continue
            
            empty_count = 0  # Reset counter
            model_name = str(header_value).strip()
            models.append({
                "name": model_name,
                "col_index": col_idx,
                "col_atchi": col_idx + 4
            })
        
        wb.Close(False)
        
    except Exception as e:
        pass
    finally:
        try:
            if 'excel' in locals():
                try:
                    # Close all open workbooks
                    while excel.Workbooks.Count > 0:
                        excel.Workbooks(1).Close(SaveChanges=False)
                except:
                    pass
                
                try:
                    excel.Visible = False  # Ensure hidden
                    excel.ScreenUpdating = True
                    excel.Quit()
                except:
                    pass
        except:
            pass
        
        try:
            pythoncom.CoUninitialize()
        except:
            pass
    
    return models

# ==========================================
# BACKGROUND WORKERS
# ==========================================

def background_worker_mode1(spek_path, selected, template_path, output_queue):
    """Worker thread untuk Mode 1 - Model Selection"""
    try:
        start_time = time.time()
        
        # Setup paths
        temp_dir = tempfile.gettempdir()
        safe_name = "".join(c for c in selected["name"] if c.isalnum() or c == ' ').strip()
        output_filename = f"Checksheet_{safe_name.replace(' ', '_')}.xlsm"
        output_path = os.path.join(temp_dir, output_filename)
        
        shutil.copy(spek_path, output_path)
        
        # Send progress updates
        output_queue.put({"progress": 10, "message": "🔌 Initializing Excel..."})
        excel = init_excel_app()
        
        output_queue.put({"progress": 20, "message": "📁 Opening workbook..."})
        wb = excel.Workbooks.Open(
            os.path.abspath(output_path),
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True
        )
        excel.Calculation = ExcelConst.CALCULATION_MANUAL
        
        output_queue.put({"progress": 30, "message": "📝 Creating dump sheet..."})
        create_dump_sheet(wb)
        
        output_queue.put({"progress": 40, "message": "📑 Injecting templates..."})
        if os.path.exists(template_path):
            inject_template_sheets(wb, template_path, excel)
        
        output_queue.put({"progress": 50, "message": "✂️ Processing sheets..."})
        process_sheet_cutting(wb, selected["col_index"], excel)
        
        output_queue.put({"progress": 80, "message": "🧹 Cleaning..."})
        delete_unused_sheets(wb)
        
        output_queue.put({"progress": 90, "message": "💾 Saving..."})
        wb.Save()
        wb.Close(SaveChanges=False)
        close_excel_safely(excel)
        
        duration = time.time() - start_time
        
        # Send completion
        output_queue.put({
            "status": "success",
            "progress": 100,
            "message": f"✅ Complete! ({duration:.1f}s)",
            "output_path": output_path,
            "duration": duration
        })
        
    except Exception as e:
        output_queue.put({
            "status": "error",
            "message": str(e),
            "traceback": traceback.format_exc()
        })
    finally:
        # Ensure Excel is properly closed even if error occurs
        try:
            if 'excel' in locals():
                try:
                    # Close all open workbooks
                    while excel.Workbooks.Count > 0:
                        excel.Workbooks(1).Close(SaveChanges=False)
                except:
                    pass
                
                try:
                    excel.Visible = False
                    excel.ScreenUpdating = True
                    excel.DisplayAlerts = True
                    excel.Quit()
                except:
                    pass
                
                try:
                    pythoncom.CoUninitialize()
                except:
                    pass
        except:
            pass


def background_worker_mode2(spek_path, checksheet_path, selected_model, output_queue):
    """Worker thread untuk Mode 2 - Model Selection with Double Validation"""
    try:
        start_time = time.time()
        
        # Detect sheet names based on checksheet filename
        checksheet_filename = os.path.basename(checksheet_path)
        checksheet_sheet_name, mib_sheet_name = detect_sheet_names(checksheet_filename)
        
        output_queue.put({"progress": 10, "message": f"🔌 Initializing Excel... (Using {mib_sheet_name})"})
        
        pythoncom.CoInitialize()
        
        try:
            excel = win32.gencache.EnsureDispatch("Excel.Application")
            excel.Visible = False  # Set immediately to prevent flash
        except:
            excel = win32.Dispatch("Excel.Application")
            excel.Visible = False  # Set immediately to prevent flash
        
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.EnableEvents = False
        
        output_queue.put({"progress": 20, "message": f"📖 Reading MIB data from {mib_sheet_name}..."})
        
        # Open Spek file
        wb_spek = excel.Workbooks.Open(os.path.abspath(spek_path), ReadOnly=True)
        ws_private = wb_spek.Worksheets(mib_sheet_name)
        
        # Find model column
        model_col_private = None
        model_col_value = None
        empty_count = 0
        for col_idx in range(11, 200, 1):  # Scan all columns
            header_value = ws_private.Cells(10, col_idx).Value
            if not header_value or str(header_value).strip() == "":
                empty_count += 1
                if empty_count >= 20:  # Stop after 20 consecutive empty columns
                    break
                continue
            
            empty_count = 0  # Reset counter
            if str(header_value).strip() == selected_model:
                model_col_private = col_idx
                model_col_value = col_idx + 4
                break
        
        if not model_col_private:
            raise Exception(f"Model '{selected_model}' not found in {mib_sheet_name}")
        
        # Read OID and Attribute mappings
        last_row_private = ws_private.Cells(ws_private.Rows.Count, 6).End(-4162).Row
        private_oid_value_map = {}
        private_attr_value_map = {}
        mib_oid_to_attr_map = {}
        parent_oid_index = {}  # NEW: Maps OID prefix → list of full OIDs for faster parent lookup
        
        if model_col_value and last_row_private >= 12:
            oid_range = ws_private.Range(ws_private.Cells(12, 6), ws_private.Cells(last_row_private, 6)).Value
            attr_range = ws_private.Range(ws_private.Cells(12, 5), ws_private.Cells(last_row_private, 5)).Value
            value_range = ws_private.Range(ws_private.Cells(12, model_col_value), ws_private.Cells(last_row_private, model_col_value)).Value
            
            if oid_range and attr_range and value_range:
                for i in range(len(oid_range)):
                    oid_row = oid_range[i]
                    attr_row = attr_range[i]
                    value_row = value_range[i]
                    
                    oid = oid_row[0] if oid_row and oid_row[0] else None
                    attr_name = attr_row[0] if attr_row and attr_row[0] else None
                    value = value_row[0] if value_row and value_row[0] else None
                    
                    if oid:
                        oid_key = str(oid).strip()
                        private_oid_value_map[oid_key] = value if value else ""
                        
                        # NEW: Build parent OID index - split by dots for prefix matching
                        oid_parts = oid_key.split('.')
                        for length in range(len(oid_parts), 0, -1):
                            prefix = '.'.join(oid_parts[:length])
                            if prefix not in parent_oid_index:
                                parent_oid_index[prefix] = []
                            if oid_key not in parent_oid_index[prefix]:
                                parent_oid_index[prefix].append(oid_key)
                        
                        if attr_name:
                            attr_normalized = str(attr_name).strip().lower()
                            mib_oid_to_attr_map[oid_key] = attr_normalized
                    
                    if attr_name:
                        attr_key = str(attr_name).strip().lower()
                        # Only map if we have a value
                        if value is not None and str(value).strip() != "":
                            private_attr_value_map[attr_key] = value
        
        # Build OID sets
        last_row_mib = ws_private.Cells(ws_private.Rows.Count, 6).End(-4162).Row
        mib_oids_list = []
        
        if last_row_mib >= 12:
            mib_range = ws_private.Range(ws_private.Cells(12, 6), ws_private.Cells(last_row_mib, 6)).Value
            if mib_range:
                mib_oids_list = [str(row[0]).strip() for row in mib_range if row[0] is not None]
        
        mib_oids_set = set(mib_oids_list)
        
        wb_spek.Close(False)
        
        output_queue.put({"progress": 40, "message": "⚙️ Processing matching..."})
        
        # Open checksheet
        wb_check = excel.Workbooks.Open(os.path.abspath(checksheet_path))
        ws_check = wb_check.Worksheets(checksheet_sheet_name)
        start_row = 10
        last_row_check = ws_check.Cells(ws_check.Rows.Count, 6).End(-4162).Row
        total_rows = last_row_check - start_row + 1
        
        # Clear old content
        if last_row_check >= start_row:
            ws_check.Range(ws_check.Cells(start_row, 9), ws_check.Cells(last_row_check, 22)).ClearContents()
        
        # Read checksheet data
        check_attr_range = []
        check_col_e_range = []
        check_oid_range = []
        
        if last_row_check >= start_row:
            check_attr_range = ws_check.Range(ws_check.Cells(start_row, 4), ws_check.Cells(last_row_check, 4)).Value
            check_col_e_range = ws_check.Range(ws_check.Cells(start_row, 5), ws_check.Cells(last_row_check, 5)).Value
            check_oid_range = ws_check.Range(ws_check.Cells(start_row, 6), ws_check.Cells(last_row_check, 6)).Value
            
            if check_attr_range is None:
                check_attr_range = []
            elif not isinstance(check_attr_range[0], tuple):
                check_attr_range = [(check_attr_range,)]
            
            if check_col_e_range is None:
                check_col_e_range = []
            elif not isinstance(check_col_e_range[0], tuple):
                check_col_e_range = [(check_col_e_range,)]
                
            if check_oid_range is None:
                check_oid_range = []
            elif not isinstance(check_oid_range[0], tuple):
                check_oid_range = [(check_oid_range,)]
        
        # Process matching with double validation
        output_data_col_i = []
        output_data_col_j = []
        output_data_rest = []
        match_count = 0
        no_support_count = 0
        
        # Statistics tracking
        value_from_exact_oid = 0
        value_from_parent_oid = 0
        value_from_attr = 0
        value_empty = 0
        
        for idx, row in enumerate(check_oid_range, start=1):
            c_oid_raw = row[0] if row else None
            
            c_attr_name = ""
            if idx <= len(check_attr_range):
                attr_row = check_attr_range[idx - 1]
                c_attr_name = str(attr_row[0]).strip().lower() if attr_row and attr_row[0] else ""
            
            col_e_value = ""
            if idx <= len(check_col_e_range):
                col_e_row = check_col_e_range[idx - 1]
                col_e_value = str(col_e_row[0]).strip() if col_e_row and col_e_row[0] else ""
            
            if not c_oid_raw:
                output_data_col_i.append("")
                output_data_col_j.append("")
                output_data_rest.append([""] * 13)
                continue
            
            c_oid = str(c_oid_raw).strip()
            
            # Check for 範囲外
            if "範囲外" in col_e_value:
                is_match = False
            else:
                is_match = False
                
                # Strategy 1: OID exact match with attribute validation
                if c_oid in mib_oids_set:
                    if c_oid in mib_oid_to_attr_map:
                        mib_attr = mib_oid_to_attr_map[c_oid]
                        if c_attr_name == mib_attr:
                            is_match = True
                    else:
                        is_match = True
                
                # Strategy 2: Improved parent OID match with progressive prefix checking
                if not is_match:
                    c_oid_parts = c_oid.split('.')
                    
                    # Try all possible parent prefixes (from longest to shortest)
                    for length in range(len(c_oid_parts) - 1, 0, -1):
                        parent_prefix = '.'.join(c_oid_parts[:length])
                        
                        # Check if this prefix exists in MIB
                        if parent_prefix in parent_oid_index:
                            # Get all MIB OIDs that match this prefix
                            matching_oids = parent_oid_index[parent_prefix]
                            
                            # Find exact match or best parent
                            for m_oid in matching_oids:
                                if c_oid.startswith(m_oid + ".") or c_oid == m_oid:
                                    # Check attribute validation if available
                                    if m_oid in mib_oid_to_attr_map:
                                        mib_attr = mib_oid_to_attr_map[m_oid]
                                        if c_attr_name == mib_attr:
                                            is_match = True
                                            break
                                    else:
                                        is_match = True
                                        break
                            
                            if is_match:
                                break
                
                # Strategy 3: Fallback to attribute-only match
                if not is_match and c_attr_name and c_attr_name in private_attr_value_map:
                    is_match = True
            
            # Get model value with IMPROVED matching (only for matched rows)
            model_value = ""
            value_source = ""
            
            if is_match:
                # Strategy 1: Exact OID match
                if c_oid in private_oid_value_map:
                    model_value = private_oid_value_map[c_oid]
                    value_source = "exact_oid"
                    if model_value and str(model_value).strip() != "":
                        value_from_exact_oid += 1
                else:
                    # Strategy 2: Find BEST (longest) parent OID match with progressive search
                    c_oid_parts = c_oid.split('.')
                    best_match_value = None
                    best_match_len = 0
                    
                    # Try progressively shorter prefixes to find best parent
                    for length in range(len(c_oid_parts) - 1, 0, -1):
                        parent_prefix = '.'.join(c_oid_parts[:length])
                        
                        # Check all MIB OIDs that could be parents
                        for m_oid in private_oid_value_map.keys():
                            # Check if checksheet OID starts with this MIB OID
                            if c_oid.startswith(m_oid + ".") or c_oid == m_oid:
                                m_oid_len = len(m_oid)
                                if m_oid_len > best_match_len:
                                    best_match_value = private_oid_value_map[m_oid]
                                    best_match_len = m_oid_len
                        
                        # Also check exact prefix match
                        if parent_prefix in private_oid_value_map:
                            prefix_len = len(parent_prefix)
                            if prefix_len > best_match_len:
                                best_match_value = private_oid_value_map[parent_prefix]
                                best_match_len = prefix_len
                        
                        # Stop if we found a match
                        if best_match_value is not None:
                            break
                    
                    if best_match_value is not None:
                        model_value = best_match_value
                        value_source = "parent_oid"
                        if model_value and str(model_value).strip() != "":
                            value_from_parent_oid += 1
                    else:
                        # Strategy 3: Fallback to attribute name match
                        if c_attr_name and c_attr_name in private_attr_value_map:
                            model_value = private_attr_value_map[c_attr_name]
                            value_source = "attribute"
                            if model_value and str(model_value).strip() != "":
                                value_from_attr += 1
                
                # Convert value to string, extract numeric value, and handle None/empty
                if model_value is None or str(model_value).strip() == "":
                    model_value = ""
                    value_empty += 1
                else:
                    model_value = extract_numeric_value(str(model_value).strip())
            
            # Build output data
            if is_match:
                match_count += 1
                output_data_col_i.append("FactoryDefault")
                output_data_col_j.append(model_value if model_value else "")
                row_data = ["", "", "", "○", "", "", "", "○", "", "", "", "○", ""]
            else:
                no_support_count += 1
                output_data_col_i.append("NoSupport")
                output_data_col_j.append("")
                row_data = ["[NA]", "", "", "-", "[NA]", "", "", "-", "[NA]", "", "", "-", ""]
            
            output_data_rest.append(row_data)
            
            # Progress update
            if idx % 100 == 0:
                progress_pct = 40 + int((idx / total_rows) * 40)
                output_queue.put({"progress": progress_pct, "message": f"⚙️ Processing {idx}/{total_rows}..."})
        
        output_queue.put({"progress": 85, "message": "💾 Writing data..."})
        
        # Write data to Excel
        if output_data_col_i and last_row_check >= start_row:
            col_i_data = [[val] for val in output_data_col_i]
            ws_check.Range(ws_check.Cells(start_row, 9), ws_check.Cells(last_row_check, 9)).Value = col_i_data
            
            col_j_data = [[val] for val in output_data_col_j]
            ws_check.Range(ws_check.Cells(start_row, 10), ws_check.Cells(last_row_check, 10)).Value = col_j_data
            
            if len(output_data_rest) > 0:
                for col_offset in range(13):
                    col_data = [[row[col_offset]] for row in output_data_rest]
                    col_index = 11 + col_offset
                    ws_check.Range(ws_check.Cells(start_row, col_index), ws_check.Cells(last_row_check, col_index)).Value = col_data
        
        # Save file
        file_name, file_ext = os.path.splitext(checksheet_path)
        new_path_check = f"{file_name}_{selected_model}_aftergenerate{file_ext}"
        
        file_format = 52 if file_ext.lower() == '.xlsm' else 51
        
        excel.Calculate()
        excel.ScreenUpdating = False
        
        wb_check.SaveAs(new_path_check, FileFormat=file_format)
        wb_check.Close(False)
        
        # Restore settings before quit
        excel.ScreenUpdating = True
        excel.EnableEvents = True
        excel.DisplayAlerts = True
        excel.Visible = False  # Ensure hidden before quit
        excel.Quit()
        
        duration = time.time() - start_time
        
        output_queue.put({
            "status": "success",
            "progress": 100,
            "message": f"✅ Complete! ({duration:.1f}s)",
            "output_path": new_path_check,
            "duration": duration,
            "stats": {
                "match_count": match_count,
                "no_support_count": no_support_count,
                "total_rows": total_rows,
                "value_from_exact_oid": value_from_exact_oid,
                "value_from_parent_oid": value_from_parent_oid,
                "value_from_attr": value_from_attr,
                "value_empty": value_empty
            }
        })
        
    except Exception as e:
        output_queue.put({
            "status": "error",
            "message": str(e),
            "traceback": traceback.format_exc()
        })
    finally:
        # Ensure Excel is properly closed even if error occurs
        try:
            if 'excel' in locals():
                try:
                    # Close all open workbooks
                    while excel.Workbooks.Count > 0:
                        excel.Workbooks(1).Close(SaveChanges=False)
                except:
                    pass
                
                try:
                    excel.Visible = False
                    excel.ScreenUpdating = True
                    excel.DisplayAlerts = True
                    excel.Quit()
                except:
                    pass
        except:
            pass
        
        try:
            pythoncom.CoUninitialize()
        except:
            pass

# ==========================================
# STREAMLIT UI
# ==========================================

st.set_page_config(
    page_title="MIB Checksheet Generator",
    page_icon="🖨️",
    layout="wide"
)

def main():
    with st.sidebar:
        st.title("🖨️ MIB Checksheet Generator")
        st.markdown("Pilih mode operasi yang diinginkan")
        st.divider()
        mode = st.radio(
            "Pilih Menu:",
            [
                "📊 Mode 1: MIB Checksheet Overall",
                "🎯 Mode 2: MIB Checksheet Section"
            ],
            index=0
        )
        
        st.divider()
        st.markdown("### ℹ️ Info")
        
        if "Mode 1" in mode:
            st.info("""
            **MIB Checksheet Overall**
            - Generate checksheet per Model
            - Fast detection dengan openpyxl
            - Template injection otomatis
            - Background processing
            """)
        else:
            st.info("""
            **MIB Checksheet Section**
            - Generate per Model
            - OID + Attribute validation
            - Prevents false positives
            - Background processing
            """)
    
    if "Mode 1" in mode:
        st.markdown("## 📊 MIB Checksheet Overall")
        st.caption("Generate checksheet dari MIB Implementation Specification dengan memilih model tertentu")
        st.divider()
        run_mode_ccode()
    else:
        st.markdown("## 🎯 MIB Checksheet Section")
        st.caption("Generate checksheet dengan matching OID + Attribute untuk model tertentu")
        st.divider()
        run_mode_model()

# ==========================================
# MODE 1: MODEL SELECTION
# ==========================================

def run_mode_ccode():
    """Mode 1: Model Selection dengan background processing""" 
    
    # Template Sheet otomatis dari folder project
    template_path = os.path.join(os.path.dirname(__file__), "Template Sheet.xlsm")
    if os.path.exists(template_path):
        st.success(f"✅ Template Sheet: Loaded automatically")
    else:
        st.warning(f"⚠️ Template Sheet not found")
    
    uploaded_spek = st.file_uploader("📂 MIB Implementation Specification (.xlsm)", type=["xlsm"], key="mode1_spek")
    
    if not uploaded_spek:
        st.info("👆 Upload MIB Implementation Specification")
        return
    
    # Save uploaded file
    temp_dir = tempfile.gettempdir()
    spek_path = os.path.join(temp_dir, uploaded_spek.name)
    
    with open(spek_path, "wb") as f:
        f.write(uploaded_spek.getbuffer())
    
    st.success(f"✅ File uploaded: {uploaded_spek.name}")
    
    # Detect Models
    st.markdown("### 🔍 Model Detection")
    
    cache_key = f"{uploaded_spek.name}_{uploaded_spek.size}"
    
    if "mode1_ccodes" not in st.session_state or st.session_state.get('mode1_cache_key') != cache_key:
        if not OPENPYXL_AVAILABLE:
            st.error("❌ openpyxl not installed. Run: pip install openpyxl")
            return
        
        with st.spinner("⚡ Detecting Models..."):
            try:
                ccodes = read_ccodes_openpyxl(spek_path)
                st.session_state.mode1_ccodes = ccodes
                st.session_state.mode1_cache_key = cache_key
                st.session_state.mode1_spek_path = spek_path
                st.success(f"⚡ Detected {len(ccodes)} Models")
            except Exception as e:
                st.error(f"❌ Detection failed: {e}")
                return
    else:
        st.info(f"📦 Using cached detection: {len(st.session_state.mode1_ccodes)} Models")
    
    if not st.session_state.mode1_ccodes:
        st.warning("⚠️ No Models detected")
        return
    
    # Model Selection & Generate
    st.markdown("### 🎯 Generate Checksheet")
    
    ccodes_list = [c["name"] for c in st.session_state.mode1_ccodes]
    selected_name = st.selectbox("Select Model:", ccodes_list, key="mode1_ccode_select")
    
    if st.button("🚀 Generate Checksheet", type="primary", use_container_width=True, key="mode1_generate"):
        selected = next(c for c in st.session_state.mode1_ccodes if c["name"] == selected_name)
        
        # Initialize queue for background worker
        result_queue = queue.Queue()
        
        # Start background thread
        worker_thread = threading.Thread(
            target=background_worker_mode1,
            args=(st.session_state.mode1_spek_path, selected, template_path, result_queue),
            daemon=True
        )
        worker_thread.start()
        
        # UI tracking
        progress_bar = st.progress(0, text="Starting...")
        
        # Poll queue for updates
        with st.status(f"⚙️ Generating {selected_name}...", expanded=True) as status:
            while worker_thread.is_alive() or not result_queue.empty():
                try:
                    update = result_queue.get(timeout=0.1)
                    
                    if "progress" in update:
                        progress_bar.progress(update["progress"], text=update["message"])
                        st.write(update["message"])
                    
                    if update.get("status") == "success":
                        status.update(label=update["message"], state="complete")
                        
                        output_path = update["output_path"]
                        if os.path.exists(output_path):
                            file_size = os.path.getsize(output_path) / (1024 * 1024)
                            st.success(f"✅ Generated: {os.path.basename(output_path)} ({file_size:.2f} MB)")
                            
                            with open(output_path, "rb") as f:
                                st.download_button(
                                    label="📥 Download Generated Checksheet",
                                    data=f,
                                    file_name=os.path.basename(output_path),
                                    mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                                    use_container_width=True
                                )
                            st.balloons()
                        break
                    
                    elif update.get("status") == "error":
                        status.update(label="❌ Error occurred!", state="error")
                        st.error(f"❌ Generation failed: {update['message']}")
                        with st.expander("🔍 Error Details"):
                            st.code(update.get("traceback", ""), language="text")
                        break
                
                except queue.Empty:
                    continue

# ==========================================
# MODE 2: MODEL SELECTION
# ==========================================

def run_mode_model():
    """Mode 2: Model selection dengan background processing"""
    
    st.markdown("### 📥 Upload Files")
    
    col1, col2 = st.columns(2)
    
    with col1:
        uploaded_spek = st.file_uploader(
            "📂 MIB Implementation Specification (.xlsm)", 
            type=["xlsm"], 
            key="mode2_spek",
            help="File PRZ-*.xlsm yang berisi EPSON Private MIB"
        )
    
    with col2:
        uploaded_checksheet = st.file_uploader(
            "📋 Checksheet Template (.xlsm)", 
            type=["xlsm"], 
            key="mode2_checksheet",
            help="File 初期値評価仕様書 (Printer Private sheet)"
        )
    
    if not uploaded_spek or not uploaded_checksheet:
        st.info("👆 Upload kedua file untuk mulai")
        return
    
    # Save uploaded files
    temp_dir = tempfile.gettempdir()
    
    spek_path = os.path.join(temp_dir, uploaded_spek.name)
    checksheet_path = os.path.join(temp_dir, uploaded_checksheet.name)
    
    with open(spek_path, "wb") as f:
        f.write(uploaded_spek.getbuffer())
    
    with open(checksheet_path, "wb") as f:
        f.write(uploaded_checksheet.getbuffer())
    
    st.success(f"✅ Files uploaded: {uploaded_spek.name}, {uploaded_checksheet.name}")
    
    # Detect models
    st.markdown("### 🔍 Model Detection")
    
    if "models_detected" not in st.session_state:
        with st.spinner("🔍 Detecting available models..."):
            try:
                checksheet_filename = os.path.basename(checksheet_path)
                models = detect_models_from_spek(spek_path, checksheet_filename)
                
                if models:
                    st.session_state.models_detected = models
                    st.session_state.spek_path = spek_path
                    st.session_state.checksheet_path = checksheet_path
                    st.success(f"✅ Detected {len(models)} models: {[m['name'] for m in models]}")
                else:
                    st.error("❌ No models detected")
                    return
            
            except Exception as e:
                st.error(f"❌ Error detecting models: {e}")
                return
    else:
        models = st.session_state.models_detected
        st.info(f"📦 {len(models)} models detected: {[m['name'] for m in models]}")
    
    # Model selection
    st.markdown("### 🎯 Generate Checksheet")
    
    model_names = [m["name"] for m in st.session_state.models_detected]
    selected_model = st.selectbox(
        "Select Model:",
        model_names,
        help="Pilih model yang ingin di-generate"
    )
    
    if st.button("🚀 Generate Checksheet", type="primary", use_container_width=True):
        result_queue = queue.Queue()
        
        worker_thread = threading.Thread(
            target=background_worker_mode2,
            args=(st.session_state.spek_path, st.session_state.checksheet_path, selected_model, result_queue),
            daemon=True
        )
        worker_thread.start()
        
        progress_bar = st.progress(0, text="Starting...")
        
        with st.status(f"⚙️ Generating {selected_model}...", expanded=True) as status:
            while worker_thread.is_alive() or not result_queue.empty():
                try:
                    update = result_queue.get(timeout=0.1)
                    
                    if "progress" in update:
                        progress_bar.progress(update["progress"], text=update["message"])
                        st.write(update["message"])
                    
                    if update.get("status") == "success":
                        status.update(label=update["message"], state="complete")
                        
                        output_path = update["output_path"]
                        if os.path.exists(output_path):
                            file_size = os.path.getsize(output_path) / (1024 * 1024)
                            st.success(f"✅ Generated: {os.path.basename(output_path)} ({file_size:.2f} MB)")
                            
                            # Show detailed stats
                            if "stats" in update:
                                stats = update["stats"]
                                col1, col2 = st.columns(2)
                                with col1:
                                    st.metric("Match", stats['match_count'], help="OID yang di-support")
                                    st.metric("NoSupport", stats['no_support_count'], help="OID tidak di-support")
                                with col2:
                                    st.metric("Total Rows", stats['total_rows'])
                                
                                # Value source breakdown
                                with st.expander("📊 Value Source Details"):
                                    st.markdown(f"""
                                    - **Exact OID Match**: {stats.get('value_from_exact_oid', 0)} values
                                    - **Parent OID Match**: {stats.get('value_from_parent_oid', 0)} values
                                    - **Attribute Match**: {stats.get('value_from_attr', 0)} values
                                    - **Empty Values**: {stats.get('value_empty', 0)} (matched but no value)
                                    """)
                            
                            with open(output_path, "rb") as f:
                                st.download_button(
                                    label="📥 Download Generated Checksheet",
                                    data=f,
                                    file_name=os.path.basename(output_path),
                                    mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                                    use_container_width=True
                                )
                            st.balloons()
                        break
                    
                    elif update.get("status") == "error":
                        status.update(label="❌ Error occurred!", state="error")
                        st.error(f"❌ Generation failed: {update['message']}")
                        with st.expander("🔍 Error Details"):
                            st.code(update.get("traceback", ""), language="text")
                        break
                
                except queue.Empty:
                    continue

if __name__ == "__main__":
    main()
